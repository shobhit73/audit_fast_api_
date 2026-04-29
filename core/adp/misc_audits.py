import pandas as pd
import io
import re
import numpy as np
import openpyxl
from datetime import datetime, date
from utils.audit_utils import norm_blank, try_parse_date, normalize_id, smart_read_df

# --- Production Logic for License Audit ---
def run_adp_license_audit(uzio_content, adp_content):
    def read_df(c, key_col):
        df = smart_read_df(c, header=None, dtype=str)
        header_idx = -1
        for i, row in df.head(20).iterrows():
            if any(str(v).strip() == key_col for v in row.values if pd.notna(v)):
                header_idx = i; break
        if header_idx != -1:
            df.columns = [str(v).strip() if pd.notna(v) else f"U_{i}" for i, v in enumerate(df.iloc[header_idx])]
            return df.iloc[header_idx+1:].reset_index(drop=True)
        return smart_read_df(c, dtype=str)

    uzio = read_df(uzio_content, "Employee ID")
    adp = read_df(adp_content, "Associate ID")
    
    UZIO_NUM, UZIO_DATE = 'License Number', 'License Expiration Date'
    ADP_NUM = 'License/Certification Code' if 'License/Certification Code' in adp.columns else 'License/Certification ID'
    ADP_DATE = 'Expiration Date'
    
    rows = []
    uzio_keys = set(uzio['Employee ID'].dropna().unique())
    adp_keys = set(adp['Associate ID'].dropna().unique())
    
    uzio_map = {k: g.to_dict('records') for k, g in uzio.groupby('Employee ID')}
    adp_map = {k: g.to_dict('records') for k, g in adp.groupby('Associate ID')}
    
    processed_adp = set()
    for eid, uz_recs in uzio_map.items():
        for uz_r in uz_recs:
            uz_n = str(uz_r.get(UZIO_NUM, "")).strip()
            if not uz_n: continue
            uz_d = try_parse_date(uz_r.get(UZIO_DATE, ""))
            
            adp_recs = adp_map.get(eid, [])
            match = next((a for a in adp_recs if str(a.get(ADP_NUM, "")).strip().lower() == uz_n.lower()), None)
            
            if match:
                processed_adp.add((eid, uz_n.lower()))
                adp_d = try_parse_date(match.get(ADP_DATE, ""))
                rows.append({"Employee ID": eid, "Field": "License Number", "Status": "Data Match", "Uzio Value": uz_n, "ADP Value": uz_n})
                rows.append({"Employee ID": eid, "Field": "Expiration Date", "Status": "Data Match" if uz_d == adp_d else "Data Mismatch", "Uzio Value": uz_d, "ADP Value": adp_d})
            else:
                rows.append({"Employee ID": eid, "Field": "License Number", "Status": "Missing in ADP", "Uzio Value": uz_n, "ADP Value": ""})

    for eid, adp_recs in adp_map.items():
        for adp_r in adp_recs:
            adp_n = str(adp_r.get(ADP_NUM, "")).strip()
            if not adp_n or (eid, adp_n.lower()) in processed_adp: continue
            rows.append({"Employee ID": eid, "Field": "License Number", "Status": "Missing in Uzio", "Uzio Value": "", "ADP Value": adp_n})

    return {"License Audit Results": rows}

# --- Production Logic for Emergency Audit ---
def run_adp_emergency_audit(uzio_content, adp_content):
    def norm_phone(x):
        s = re.sub(r"\D", "", str(x or ""))
        if len(s) == 11 and s.startswith("1"): s = s[1:]
        return s

    uzio = smart_read_df(uzio_content, header=1, dtype=str)
    adp = smart_read_df(adp_content, dtype=str)
    
    u_eid, u_name, u_rel, u_ph = 'Employee ID', 'Name', 'Relationship', 'Phone'
    a_eid = next((c for c in adp.columns if "Associate ID" in c), "Associate ID")
    a_name = next((c for c in adp.columns if "Contact Name" in c), "Contact Name")
    a_rel = next((c for c in adp.columns if "Relationship Description" in c), "Relationship Description")
    a_ph = next((c for c in adp.columns if "Mobile Phone" in c), "Mobile Phone")

    rows = []
    u_map = {k: g.to_dict('records') for k, g in uzio.groupby(u_eid)}
    a_map = {k: g.to_dict('records') for k, g in adp.groupby(a_eid)}
    
    all_ids = set(u_map.keys()) | set(a_map.keys())
    for eid in sorted(all_ids):
        u_recs, a_recs = u_map.get(eid, []), a_map.get(eid, [])
        for u in u_recs:
            u_n, u_p = str(u.get(u_name, "")).strip().lower(), norm_phone(u.get(u_ph, ""))
            match = next((a for a in a_recs if str(a.get(a_name, "")).strip().lower() == u_n or norm_phone(a.get(a_ph, "")) == u_p), None)
            status = "Data Match" if match else "Missing in ADP"
            rows.append({"Employee ID": eid, "Field": "Contact Name", "Status": status, "Uzio Value": u.get(u_name), "ADP Value": match.get(a_name) if match else ""})
        for a in a_recs:
            if not any(str(u.get(u_name, "")).strip().lower() == str(a.get(a_name, "")).strip().lower() for u in u_recs):
                rows.append({"Employee ID": eid, "Field": "Contact Name", "Status": "Missing in Uzio", "Uzio Value": "", "ADP Value": a.get(a_name)})

    df_res = pd.DataFrame(rows)
    summary = []
    if not df_res.empty:
        counts = df_res["Status"].value_counts().reset_index()
        counts.columns = ["Status", "Count"]
        summary = counts.to_dict(orient="records")
    return {"Emergency_Contact_Audit": rows, "Summary": summary}

# --- Production Logic for Timeoff Audit ---
def run_adp_timeoff_audit(uzio_content, adp_content):
    adp = pd.read_excel(io.BytesIO(adp_content))
    a_id = next((c for c in adp.columns if "ASSOCIATE ID" in c.upper()), None)
    a_bal = next((c for c in adp.columns if "BALANCE AMOUNT" in c.upper()), None)
    
    if not a_id or not a_bal: return {"error": "ADP file missing ID or Balance columns"}
    
    adp['Clean_ID'] = adp[a_id].apply(lambda x: str(x).strip().lstrip("0").replace(".0", ""))
    balance_map = adp.groupby('Clean_ID')[a_bal].sum().to_dict()
    
    wb = openpyxl.load_workbook(io.BytesIO(uzio_content))
    ws = wb.worksheets[1] # Time Off Details
    
    # Update logic from tool
    header_row = 4
    idx_id, idx_bal = None, None
    for cell in ws[header_row]:
        v = str(cell.value).strip() if cell.value else ""
        if "Employee ID" in v: idx_id = cell.column
        elif "Opening Balance" in v: idx_bal = cell.column
    
    updates = 0
    if idx_id and idx_bal:
        for r in range(header_row + 1, ws.max_row + 1):
            eid = str(ws.cell(row=r, column=idx_id).value or "").strip().lstrip("0").replace(".0", "")
            if eid in balance_map:
                ws.cell(row=r, column=idx_bal).value = balance_map[eid]
                updates += 1

    # Return summary of updates
    return {"message": f"Updated {updates} balances in Uzio template", "status": "Success"}
