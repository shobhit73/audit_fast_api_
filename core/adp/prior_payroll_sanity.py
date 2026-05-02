"""ADP - Prior Payroll Sanity Check (MCP core).

Pure Python port of the Streamlit `apps/adp/prior_payroll_sanity.py` tool.
Cleans an ADP Prior Payroll file before downstream API ingestion:

  1. Drops the interleaved 'Totals For Associate ID XYZ:' summary rows.
  2. Detects and removes the bottom-of-file grand-total row (where the
     last employee's ID got bled into the totals row).
  3. Auto-detects per-pay-period exports (multiple rows per associate)
     and aggregates them into one row per associate using SUM for
     money/hours, MIN/MAX for period dates, and first-non-null for
     identity columns. Same-pay-date duplicates (real distinct paychecks
     in ADP) are also folded together by SUM, which is correct for ADP.
  4. Optionally swaps NET PAY <-> TAKE HOME values (the Carvan API maps
     them reversed). Column headers are NEVER changed -- only the data.

Output is CSV bytes with the input's exact column headers and order.
The pipeline returns (csv_bytes, summary_dict).
"""

import re
import io
import pandas as pd
import openpyxl
from utils.audit_utils import clean_money_val


def _find_col(df, candidates):
    """Case-insensitive exact-then-substring lookup of a column."""
    for cand in candidates:
        for c in df.columns:
            if str(c).strip().lower() == cand.lower():
                return c
    for cand in candidates:
        for c in df.columns:
            if cand.lower() in str(c).strip().lower():
                return c
    return None


_ROUND_FORMULA_RE = re.compile(r"^=ROUND\(\s*(-?[\d.]+)\s*,\s*[\d.]+\s*\)\s*$", re.IGNORECASE)


def _evaluate_cell(value):
    """Resolve =ROUND(x,n) formulas (the only formula style ADP exports use for money cells)."""
    if value is None:
        return None
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s.startswith("="):
        return value
    m = _ROUND_FORMULA_RE.match(s)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def _read_excel_with_formula_eval(buf):
    """Read .xlsx/.xls bytes via openpyxl, evaluating =ROUND() formulas."""
    wb = openpyxl.load_workbook(buf, data_only=False)
    target_sheet = wb.sheetnames[0]
    if len(wb.sheetnames) > 1 and "criteria" in wb.sheetnames[0].lower():
        target_sheet = wb.sheetnames[1]
    ws = wb[target_sheet]

    header_idx = 0
    for r in range(1, min(ws.max_row, 50) + 1):
        row_text = " ".join(
            str(ws.cell(r, c).value).lower()
            for c in range(1, ws.max_column + 1)
            if ws.cell(r, c).value is not None
        )
        if any(k in row_text for k in ["associate id", "employee id", "file #"]):
            header_idx = r - 1
            break

    headers = [ws.cell(header_idx + 1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(header_idx + 2, ws.max_row + 1):
        row = [_evaluate_cell(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        rows.append(row)
    return pd.DataFrame(rows, columns=headers), header_idx, target_sheet


def read_input_bytes(content, filename):
    """Read an ADP file from raw bytes. Dispatches to CSV or Excel reader by filename suffix."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        buf = io.BytesIO(content)
        df_peek = pd.read_csv(buf, header=None, nrows=50, dtype=str)
        header_idx = 0
        for i, row in df_peek.iterrows():
            row_str = " ".join(str(x).lower() for x in row if pd.notna(x))
            if any(k in row_str for k in ["associate id", "employee id", "file #"]):
                header_idx = i
                break
        buf.seek(0)
        df = pd.read_csv(buf, header=header_idx, dtype=str)
        return df, header_idx, "Sheet1"
    return _read_excel_with_formula_eval(io.BytesIO(content))


def drop_summary_rows(df):
    """Drop 'Totals For Associate ID' summary rows where Associate ID is null."""
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    if not eid_col:
        return df.reset_index(drop=True), 0
    mask = df[eid_col].notna() & (df[eid_col].astype(str).str.strip() != "")
    removed = (~mask).sum()
    return df[mask].reset_index(drop=True), int(removed)


def _smart_merge_value(values):
    """Pick the best value across duplicate rows for a single column.
    Used by merge_duplicate_pay_periods (the 'preserve_pay_periods' strategy).
    """
    cleaned = []
    for v in values:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        sv = str(v).strip()
        if sv in ("", "-", "nan", "NaT"):
            continue
        cleaned.append(v)
    if not cleaned:
        return values[0] if len(values) > 0 else None

    best_num = None
    best_num_val = None
    for v in cleaned:
        try:
            num = clean_money_val(v)
            if best_num is None or abs(num) > abs(best_num_val):
                best_num = v
                best_num_val = num
        except Exception:
            continue
    if best_num is not None and best_num_val is not None and best_num_val != 0:
        return best_num
    return cleaned[0]


def merge_duplicate_pay_periods(df):
    """Fold rows that share Employee ID + Pay Date (+ optional period start/end) into
    one row, but keep distinct pay periods intact. Used by the 'preserve_pay_periods'
    aggregation strategy. Returns (cleaned_df, list_of_merge_events).
    """
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    pay_col = _find_col(df, ["Pay Date", "Check Date", "Pay Period End Date"])
    start_col = _find_col(df, ["Period Start", "Pay Period Start", "Start Date"])
    end_col = _find_col(df, ["Period End", "Pay Period End", "End Date"])

    if not eid_col or not pay_col:
        return df, []

    keys = [eid_col, pay_col]
    if start_col and start_col not in keys:
        keys.append(start_col)
    if end_col and end_col not in keys:
        keys.append(end_col)

    work = df.copy()
    work["_orig_idx"] = range(len(work))

    grouped = work.groupby(keys, dropna=False, sort=False)
    counts = grouped.size().reset_index(name="_n")
    dup_keys = counts[counts["_n"] > 1]
    if dup_keys.empty:
        return df.reset_index(drop=True), []

    drop_indices = set()
    merge_events = []
    merged_records = []

    for key_vals, group in grouped:
        if len(group) == 1:
            continue
        first_idx = int(group["_orig_idx"].iloc[0])
        merged = {col: _smart_merge_value(group[col].tolist()) for col in df.columns}
        merged_records.append((first_idx, merged))
        drop_indices.update(int(i) for i in group["_orig_idx"].tolist())
        merge_events.append({
            "Employee ID": str(key_vals[0]),
            "Pay Date": str(key_vals[1]),
            "Rows merged": int(len(group)),
            "Kept canonical row at original index": first_idx,
        })

    cleaned_rows = []
    for i in range(len(df)):
        if i in drop_indices:
            continue
        cleaned_rows.append(df.iloc[i].to_dict())
    for first_idx, merged in merged_records:
        merged["_insert_at"] = first_idx
        cleaned_rows.append(merged)
    cleaned_rows.sort(key=lambda r: r.get("_insert_at", -1) if "_insert_at" in r else -1)
    for r in cleaned_rows:
        r.pop("_insert_at", None)

    cleaned = pd.DataFrame(cleaned_rows, columns=df.columns)
    return cleaned.reset_index(drop=True), merge_events


def detect_per_pay_period_structure(df):
    """Return ('aggregate', summary) when any associate has more than one row,
    otherwise ('none', summary). Aggregation is the right move for ADP files where
    same-pay-date duplicates are real distinct paychecks (different check numbers)
    and per-pay-period rows are the implementor's most common export mistake.
    """
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    pay_col = _find_col(df, ["Pay Date", "Check Date"])
    if not eid_col:
        return "none", None
    work = df[df[eid_col].notna()].copy()
    work[eid_col] = work[eid_col].astype(str).str.strip()
    work = work[work[eid_col] != ""]
    if work.empty:
        return "none", None
    rows_per_eid = work.groupby(eid_col).size()
    summary = {
        "associates": int(len(rows_per_eid)),
        "with_multiple_rows": int((rows_per_eid > 1).sum()),
        "max_rows_for_single_associate": int(rows_per_eid.max()),
    }
    if pay_col:
        pay_dates_per_eid = work.groupby(eid_col)[pay_col].nunique()
        summary["max_pay_dates_for_single_associate"] = int(pay_dates_per_eid.max())
    return ("aggregate" if summary["with_multiple_rows"] > 0 else "none"), summary


def _to_float(v):
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", "-", "nan", "NaT"):
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def _format_date(dt):
    if pd.isna(dt):
        return None
    try:
        return dt.strftime("%m/%d/%Y")
    except Exception:
        return None


def aggregate_by_associate(df):
    """Aggregate per-pay-period rows into one row per Associate ID."""
    eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
    pay_col = _find_col(df, ["Pay Date", "Check Date"])
    period_begin_col = _find_col(df, ["Period Beginning Date", "Period Begin Date", "Start Date"])
    period_end_col = _find_col(df, ["Period Ending Date", "Period End Date", "End Date"])
    term_col = _find_col(df, ["Termination Date"])
    check_col = _find_col(df, ["Check/Voucher Number", "Check Number", "Voucher Number"])

    min_date_cols = {period_begin_col} - {None}
    max_date_cols = {period_end_col, pay_col, term_col} - {None}
    identity_col_names = ["Name", "File Number", "Position ID", "Status", "Tax ID",
                          "Dist #", "Worked In State"]
    identity_cols = {_find_col(df, [n]) for n in identity_col_names} - {None}

    if not eid_col:
        return df, None

    EMPTY_PLACEHOLDER = "-"

    aggregated_rows = []
    for eid_val, group in df.groupby(eid_col, sort=False):
        out_row = {}
        for col in df.columns:
            vals = group[col].tolist()
            if col == eid_col:
                out_row[col] = eid_val
                continue
            if col == check_col:
                out_row[col] = ""
                continue
            if col in min_date_cols or col in max_date_cols:
                dts = pd.to_datetime(vals, errors="coerce")
                dts = dts.dropna() if hasattr(dts, "dropna") else pd.Series(dts).dropna()
                if len(dts) == 0:
                    out_row[col] = EMPTY_PLACEHOLDER
                else:
                    target = dts.min() if col in min_date_cols else dts.max()
                    out_row[col] = _format_date(target) or EMPTY_PLACEHOLDER
                continue
            if col in identity_cols:
                first = next(
                    (v for v in vals
                     if v is not None
                     and not (isinstance(v, float) and pd.isna(v))
                     and str(v).strip() not in ("", "nan", "NaT")),
                    None,
                )
                out_row[col] = first if first is not None else EMPTY_PLACEHOLDER
                continue

            numeric_vals = []
            categorical_vals = []
            for v in vals:
                f = _to_float(v)
                if f is not None:
                    numeric_vals.append(f)
                elif v is not None and not (isinstance(v, float) and pd.isna(v)):
                    s = str(v).strip()
                    if s and s.lower() not in ("nan", "nat"):
                        categorical_vals.append(v)
            if numeric_vals and not categorical_vals:
                if all(v == 0 for v in numeric_vals):
                    out_row[col] = EMPTY_PLACEHOLDER
                else:
                    out_row[col] = round(sum(numeric_vals), 2)
            elif categorical_vals:
                out_row[col] = categorical_vals[0]
            else:
                out_row[col] = EMPTY_PLACEHOLDER
        aggregated_rows.append(out_row)

    out_df = pd.DataFrame(aggregated_rows, columns=df.columns)
    return out_df, {
        "input_rows": int(len(df)),
        "output_rows": int(len(out_df)),
        "associates": int(len(out_df)),
    }


def apply_net_take_swap(df):
    """Swap NET PAY <-> TAKE HOME column values without renaming headers."""
    net_col = _find_col(df, ["Net Pay"])
    take_col = _find_col(df, ["Take Home"])
    if not net_col or not take_col or net_col == take_col:
        return df, False
    net_vals = df[net_col].copy()
    df[net_col] = df[take_col].copy()
    df[take_col] = net_vals
    return df, True


def detect_grand_total_row(df):
    """Detect & drop the bottom-of-file grand-total row where the last employee's
    ID leaked into the totals row. Returns (cleaned_df, info_dict_or_None).
    """
    if len(df) < 2:
        return df, None
    last_row = df.iloc[-1]
    prev_row = df.iloc[-2]
    shared = 0
    for c in df.columns[:5]:
        v_l = str(last_row[c]).strip()
        v_p = str(prev_row[c]).strip()
        if v_l and v_l == v_p and v_l.lower() != "nan":
            shared += 1
    if shared < 1:
        return df, None
    for c in df.columns:
        try:
            val_last = clean_money_val(last_row[c])
            if val_last <= 100:
                continue
            sum_rest = sum(clean_money_val(x) for x in df[c].iloc[:-1])
            if sum_rest > 0 and abs(val_last - sum_rest) < sum_rest * 0.05:
                eid_col = _find_col(df, ["Associate ID", "Employee ID", "File #"])
                first_col = _find_col(df, ["First Name"])
                last_col = _find_col(df, ["Last Name"])
                preview_eid = str(last_row[eid_col]) if eid_col else ""
                fn = str(last_row[first_col]).strip() if first_col and pd.notna(last_row[first_col]) else ""
                ln = str(last_row[last_col]).strip() if last_col and pd.notna(last_row[last_col]) else ""
                return df.iloc[:-1].copy(), {
                    "removed_employee_id": preview_eid,
                    "removed_employee_name": (fn + " " + ln).strip(),
                    "matched_on_column": str(c),
                    "matched_value": round(val_last, 2),
                    "expected_sum": round(sum_rest, 2),
                }
        except Exception:
            continue
    return df, None


def run_adp_prior_payroll_sanity(
    content,
    filename="upload.xlsx",
    swap_net_take=True,
    aggregation_strategy="full_quarter",
):
    """Run the full sanity-check pipeline on ADP file bytes.

    aggregation_strategy:
      - 'full_quarter' (default): when the file has multiple rows per associate,
        roll everything up into ONE row per associate (the user's "Full Quarter
        (Default)" UI option). This is the standard implementor-error-fixer.
      - 'preserve_pay_periods': keep distinct pay-period rows but merge any
        same-day duplicates within a single pay date (the user's "Preserve Pay
        Periods" UI option). Useful when the API expects per-period rows.

    Returns (csv_bytes, summary_dict). csv_bytes are UTF-8 encoded.
    """
    df_in, header_idx, sheet = read_input_bytes(content, filename)
    original_count = len(df_in)
    df_a, summary_removed = drop_summary_rows(df_in)
    df_b, gt_info = detect_grand_total_row(df_a)
    mode, period_info = detect_per_pay_period_structure(df_b)
    agg_info = None
    merge_events = None
    if mode == "aggregate":
        if str(aggregation_strategy).lower() in ("preserve_pay_periods", "preserve"):
            df_c, merge_events = merge_duplicate_pay_periods(df_b)
            mode = "preserve"
        else:
            df_c, agg_info = aggregate_by_associate(df_b)
    else:
        df_c = df_b
    swapped = False
    if swap_net_take:
        df_c, swapped = apply_net_take_swap(df_c)

    buf = io.StringIO()
    df_c.to_csv(buf, index=False)

    summary = {
        "input_rows": original_count,
        "summary_rows_removed": summary_removed,
        "grand_total_removed": gt_info is not None,
        "grand_total_info": gt_info,
        "aggregation_strategy": aggregation_strategy,
        "mode": mode,
        "period_info": period_info,
        "aggregation_info": agg_info,
        "merge_events_count": len(merge_events) if merge_events is not None else 0,
        "swap_applied": swapped,
        "output_rows": len(df_c),
        "sheet_used": sheet,
        "header_row_index": header_idx,
    }
    return buf.getvalue().encode("utf-8"), summary
