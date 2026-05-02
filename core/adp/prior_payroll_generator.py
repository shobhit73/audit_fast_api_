"""ADP -> Uzio Prior Payroll Generator (MCP core).

Pure-Python port of the Streamlit `apps/adp/prior_payroll_generator.py` tool.
Takes a blank Uzio Prior Payroll Template + 1-10 ADP Prior Payroll History files
and emits a filled Uzio template (.xlsx bytes).

Mapping: each ADP dynamic column is auto-mapped to a Uzio target column using a
fuzzy-string heuristic (auto_guess_mapping). Callers can pass override_mapping
to force specific (adp_col -> uzio_col_idx) pairs.

Output is the filled Uzio template as xlsx bytes plus a summary dict containing:
  - file_summaries: per-input-file pay period / employee counts
  - mapping: {adp_col: uzio_col_idx} actually applied
  - skipped_columns: ADP columns with non-zero values that were not mapped
  - validation_issues: employee-periods where Gross - Taxes - Deductions != Net Pay
  - output_rows / unique_employees / pay_periods
"""

import io
import re
import difflib
from collections import defaultdict
from datetime import datetime

import pandas as pd
import openpyxl


# Uzio template constants -- mirror the Streamlit version exactly
UZIO_HEADER_ROW = 5
UZIO_SECTION_ROW = 4
UZIO_DATA_START_ROW = 6
UZIO_EMPLOYEE_ID_COL = 1
UZIO_FULL_NAME_COL = 2
UZIO_SSN_COL = 3
UZIO_PP_START_COL = 4
UZIO_PP_END_COL = 5
UZIO_PAYCHECK_DATE_COL = 6
UZIO_FIRST_DATA_COL = 7

SKIP_LABEL = "__SKIP__"


def get_adp_category(col_name):
    """Categorize an ADP dynamic column."""
    name = str(col_name).strip().upper()
    if ("MEMO" in name or
        "DIRECT DEPOSIT" in name or
        name in ("GROSS PAY", "TAKE HOME", "NET PAY") or
        name.startswith("TOTAL ")):
        return "_SKIP"
    if "EARNINGS" in name or "HOURS" in name:
        return "Earnings"
    if "EMPLOYEE TAX" in name:
        return "Employee Taxes"
    if "EMPLOYER TAX" in name:
        return "Employer Taxes"
    if "DEDUCTION" in name:
        return "Deductions"
    return "Deductions"


def auto_guess_mapping(adp_col, uzio_col_headers):
    """Return the best-matching Uzio column index for the given ADP column,
    or None if no candidate scores above the threshold.
    """
    if not adp_col or not isinstance(adp_col, str):
        return None
    td_lower = adp_col.lower()
    td_clean = re.sub(r"[^a-z0-9]", "", td_lower)
    if not td_clean:
        return None

    best_idx = None
    best_score = 0.0
    for col_idx, hdr in uzio_col_headers.items():
        if col_idx < UZIO_FIRST_DATA_COL:
            continue
        hdr_lower = str(hdr).lower()
        hdr_clean = re.sub(r"[^a-z0-9]", "", hdr_lower)
        if not hdr_clean:
            continue
        score = difflib.SequenceMatcher(None, td_clean, hdr_clean).ratio()
        td_words = set(re.findall(r"[a-z0-9]+", td_lower))
        hdr_words = set(re.findall(r"[a-z0-9]+", hdr_lower))
        overlap = td_words & hdr_words
        if overlap:
            score += 0.15 * len(overlap)
        if "medicare" in td_words and "medicare" in hdr_words: score += 0.3
        if ("soc" in td_words or "ss" in td_words) and "social" in hdr_words: score += 0.3
        if ("fit" in td_words or "fed" in td_words or "federal" in td_words) and "federal" in hdr_words: score += 0.3
        if "401k" in td_lower and "401k" in hdr_lower: score += 0.3
        if "regular" in td_words and "regular" in hdr_words: score += 0.3
        if "overtime" in td_words and "overtime" in hdr_words: score += 0.3
        if "bonus" in td_words and "bonus" in hdr_words: score += 0.3
        if "futa" in td_words and "futa" in hdr_words: score += 0.3
        if "sui" in td_words and "sui" in hdr_words: score += 0.3
        if "sdi" in td_words and "sdi" in hdr_words: score += 0.3
        if "worked in state" in td_lower and "state income" in hdr_lower: score += 0.5
        if score > best_score and score >= 0.65:
            best_score = score
            best_idx = col_idx
    return best_idx


def parse_date(date_str):
    if pd.isna(date_str) or not str(date_str).strip():
        return ""
    try:
        if isinstance(date_str, datetime):
            return date_str.strftime("%m/%d/%Y")
        return pd.to_datetime(date_str).strftime("%m/%d/%Y")
    except Exception:
        return str(date_str).strip()


def read_uzio_template(content):
    """Open the blank Uzio template and pull the section + column header rows.
    Returns (section_headers, column_headers, wb, ws). The wb/ws are mutated in
    place by write_output_excel.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[-1]]
    for sname in wb.sheetnames:
        if "payroll" in sname.lower() and "instruction" not in sname.lower():
            ws = wb[sname]
            break
    section_headers = {
        cell.column: str(cell.value).strip()
        for cell in ws[UZIO_SECTION_ROW] if cell.value
    }
    column_headers = {
        cell.column: str(cell.value).strip()
        for cell in ws[UZIO_HEADER_ROW] if cell.value
    }
    return section_headers, column_headers, wb, ws


def read_adp_files(adp_files_data):
    """Read 1-10 ADP files. adp_files_data is list of (content_bytes, filename).
    Returns (all_rows, dynamic_cols, file_summaries).
    """
    all_rows = []
    all_dynamic_cols = set()
    file_summaries = []

    for content, filename in adp_files_data:
        try:
            xl = pd.ExcelFile(io.BytesIO(content))
            sheet = xl.sheet_names[0]
            df_test = xl.parse(sheet, nrows=10)
            header_row_idx = None
            for i in range(len(df_test)):
                row_vals = [str(x).upper().strip() for x in df_test.iloc[i].tolist()]
                if "FILE NUMBER" in row_vals or "COMPANY CODE" in row_vals:
                    header_row_idx = i
                    break
            df = xl.parse(sheet, header=header_row_idx + 1) if header_row_idx is not None else xl.parse(sheet)

            if "FILE NUMBER" in df.columns:
                df = df[df["FILE NUMBER"].notna()]
                df = df[~df["FILE NUMBER"].astype(str).str.contains("Total", case=False, na=False)]
            elif "COMPANY CODE" in df.columns:
                df = df[df["COMPANY CODE"].notna()]
                df = df[~df["COMPANY CODE"].astype(str).str.contains("Total", case=False, na=False)]
            if "NAME" in df.columns:
                df = df[df["NAME"].notna()]

            standard_cols = [
                "COMPANY CODE", "NAME", "FILE NUMBER", "POSITION ID", "STATUS",
                "TAX ID", "ASSOCIATE ID", "WORKED IN STATE", "DIST #",
                "PERIOD BEGINNING DATE", "PERIOD ENDING DATE", "PAY DATE",
                "CHECK/VOUCHER NUMBER",
            ]
            file_dynamic_cols = [c for c in df.columns if c not in standard_cols and not str(c).startswith("Unnamed:")]
            all_dynamic_cols.update(file_dynamic_cols)

            if not df.empty:
                min_start = df["PERIOD BEGINNING DATE"].min() if "PERIOD BEGINNING DATE" in df.columns else ""
                max_end = df["PERIOD ENDING DATE"].max() if "PERIOD ENDING DATE" in df.columns else ""
                pay_date = df["PAY DATE"].iloc[0] if "PAY DATE" in df.columns else ""
                file_summaries.append({
                    "Filename": filename,
                    "Pay Period": f"{parse_date(min_start)} - {parse_date(max_end)}",
                    "Pay Date": parse_date(pay_date),
                    "Employees": int(df["FILE NUMBER"].nunique()) if "FILE NUMBER" in df.columns else 0,
                    "Records": int(len(df)),
                })

            for _, row in df.iterrows():
                all_rows.append({k: (v if pd.notna(v) else None) for k, v in row.to_dict().items()})
        except Exception as e:
            file_summaries.append({"Filename": filename, "Error": str(e)})

    return all_rows, sorted(all_dynamic_cols), file_summaries


def generate_output(adp_rows, mapping, net_pay_col_idx):
    """Aggregate per (employee, pay-period-start) into one row per Uzio output."""
    output_rows = []
    skipped_items = []
    validation_results = []

    ee_groups = defaultdict(list)
    for r in adp_rows:
        ee_code = str(r.get("FILE NUMBER", r.get("ASSOCIATE ID", ""))).strip()
        pp_start = parse_date(r.get("PERIOD BEGINNING DATE"))
        if ee_code:
            ee_groups[(ee_code, pp_start)].append(r)

    for (ee_code, pp_start), rows in ee_groups.items():
        base_row = rows[0]
        out_row = {
            UZIO_EMPLOYEE_ID_COL: ee_code,
            UZIO_FULL_NAME_COL: str(base_row.get("NAME", "") or "").strip(),
            UZIO_PP_START_COL: pp_start,
            UZIO_PP_END_COL: parse_date(base_row.get("PERIOD ENDING DATE")),
            UZIO_PAYCHECK_DATE_COL: parse_date(base_row.get("PAY DATE")),
        }

        net_pay_total = 0.0
        gross_earnings = 0.0
        total_ee_taxes = 0.0
        total_deductions = 0.0
        total_er_taxes = 0.0

        for row in rows:
            try:
                net_amt = float(row.get("NET PAY", 0)) if pd.notna(row.get("NET PAY")) else 0.0
            except Exception:
                net_amt = 0.0
            net_pay_total += net_amt

            for adp_col, target_col in mapping.items():
                try:
                    amt = float(row.get(adp_col, 0)) if pd.notna(row.get(adp_col)) else 0.0
                except Exception:
                    amt = 0.0
                if amt == 0:
                    continue
                out_row[target_col] = out_row.get(target_col, 0) + amt
                cat = get_adp_category(adp_col)
                if cat == "Earnings":
                    gross_earnings += amt
                elif cat == "Employee Taxes":
                    total_ee_taxes += amt
                elif cat == "Employer Taxes":
                    total_er_taxes += amt
                elif cat in ("Deductions", "Contributions"):
                    total_deductions += amt

            for col in row.keys():
                if col not in mapping and get_adp_category(col) != "_SKIP":
                    try:
                        amt = float(row.get(col, 0)) if pd.notna(row.get(col)) else 0.0
                    except Exception:
                        amt = 0.0
                    if amt != 0:
                        skipped_items.append({
                            "Employee ID": ee_code,
                            "Pay Period Start": pp_start,
                            "ADP Column": col,
                            "Amount": round(amt, 2),
                        })

        if net_pay_col_idx:
            out_row[net_pay_col_idx] = net_pay_total

        expected_net = gross_earnings - total_ee_taxes - total_deductions
        if abs(expected_net - net_pay_total) > 0.02:
            validation_results.append({
                "Employee ID": ee_code,
                "Pay Period": f"{pp_start} - {out_row[UZIO_PP_END_COL]}",
                "Pushed Gross Earnings": round(gross_earnings, 2),
                "Pushed EE Taxes": round(total_ee_taxes, 2),
                "Pushed Deductions": round(total_deductions, 2),
                "Expected Net": round(expected_net, 2),
                "Actual Source Net Pay": round(net_pay_total, 2),
                "Difference": round(expected_net - net_pay_total, 2),
            })

        output_rows.append(out_row)

    output_rows.sort(key=lambda r: (str(r.get(UZIO_EMPLOYEE_ID_COL, "")), str(r.get(UZIO_PP_START_COL, ""))))
    return output_rows, skipped_items, validation_results


def write_output_excel(uzio_wb, uzio_ws, output_rows, uzio_col_headers):
    """Mutate the Uzio worksheet in place and return the workbook as xlsx bytes."""
    if uzio_ws.max_row >= UZIO_DATA_START_ROW:
        uzio_ws.delete_rows(UZIO_DATA_START_ROW, uzio_ws.max_row - UZIO_DATA_START_ROW + 1)
    max_col = max(uzio_col_headers.keys()) if uzio_col_headers else 86
    for row_idx, out_row in enumerate(output_rows):
        excel_row = UZIO_DATA_START_ROW + row_idx
        for col_idx in range(1, max_col + 1):
            val = out_row.get(col_idx)
            if val is not None:
                uzio_ws.cell(row=excel_row, column=col_idx, value=val)
    buf = io.BytesIO()
    uzio_wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def run_adp_prior_payroll_generator(uzio_template_bytes, adp_files_data, override_mapping=None):
    """End-to-end pipeline. Returns (filled_xlsx_bytes, summary_dict).

    override_mapping: optional {adp_col_name: uzio_col_idx_or_-1_to_skip} that wins
    over the auto-guessed mapping. Use -1 (or any negative) to force-skip an item.
    """
    section_headers, uzio_col_headers, uzio_wb, uzio_ws = read_uzio_template(uzio_template_bytes)
    adp_rows, adp_dynamic_cols, file_summaries = read_adp_files(adp_files_data)
    if not adp_rows:
        raise ValueError("No valid data rows found in ADP source files")

    mapping = {}
    for col in adp_dynamic_cols:
        if get_adp_category(col) == "_SKIP":
            continue
        guess = auto_guess_mapping(col, uzio_col_headers)
        if guess is not None:
            mapping[col] = guess
    if override_mapping:
        for col, target in override_mapping.items():
            if isinstance(target, (int, float)) and int(target) >= UZIO_FIRST_DATA_COL:
                mapping[col] = int(target)
            else:
                mapping.pop(col, None)

    net_pay_col_idx = None
    for col_idx, hdr in uzio_col_headers.items():
        if "net pay" in str(hdr).lower():
            net_pay_col_idx = col_idx
            break

    output_rows, skipped_items, validation_results = generate_output(
        adp_rows, mapping, net_pay_col_idx,
    )

    xlsx_bytes = write_output_excel(uzio_wb, uzio_ws, output_rows, uzio_col_headers)

    summary = {
        "input_file_summaries": file_summaries,
        "mapping": {col: int(idx) for col, idx in mapping.items()},
        "mapping_count": len(mapping),
        "skipped_with_values": skipped_items[:200],
        "skipped_count": len(skipped_items),
        "validation_issues": validation_results[:200],
        "validation_issue_count": len(validation_results),
        "output_rows": len(output_rows),
        "unique_employees": len({r.get(UZIO_EMPLOYEE_ID_COL) for r in output_rows}),
        "pay_periods": len({(r.get(UZIO_PP_START_COL), r.get(UZIO_PP_END_COL)) for r in output_rows}),
        "net_pay_target_col": net_pay_col_idx,
    }
    return xlsx_bytes, summary
