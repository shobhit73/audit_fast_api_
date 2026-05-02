"""Paycom -> Uzio Prior Payroll Generator (MCP core).

Pure-Python port of the Streamlit `apps/paycom/prior_payroll_generator.py` tool.
Reads a blank Uzio Prior Payroll Template + 1-10 Paycom Prior Payroll files
(long format with Type Code / Type Description / Code Description / Amount rows)
and emits a filled Uzio template (.xlsx bytes).

Mapping key is (type_code, type_description); each is auto-mapped to a Uzio
target column via a fuzzy-string heuristic (auto_guess_mapping). Callers can
pass override_mapping to force specific mappings. Net Pay Distribution rows are
auto-summed to the Uzio 'Net Pay' column; Employee Benefits rows are skipped.
"""

import io
import re
import difflib
from collections import defaultdict

import openpyxl


# Uzio template constants (mirror of streamlit version)
UZIO_HEADER_ROW = 5
UZIO_SECTION_ROW = 4
UZIO_DATA_START_ROW = 6
UZIO_EMPLOYEE_ID_COL = 1
UZIO_FULL_NAME_COL = 2
UZIO_SSN_COL = 3
UZIO_PP_START_COL = 4
UZIO_PP_END_COL = 5
UZIO_PAYCHECK_DATE_COL = 6
UZIO_FIRST_DATA_COL = 7

# Paycom Code Description -> UI section
PAYCOM_CATEGORY_TO_SECTION = {
    "Earnings":                "Earnings",
    "W/H Taxes":               "Employee Taxes",
    "Client Side Liabilities": "Employer Taxes",
    "Deductions":              "Deductions",
    "Net Pay Distribution":    "_NET_PAY",
    "Employee Benefits":       "_BENEFITS",
}


def auto_guess_mapping(td, cd, uzio_col_headers):
    """Best-matching Uzio column index for a Paycom Type Description, or None."""
    if not td or not isinstance(td, str):
        return None
    td_lower = td.lower()
    td_clean = re.sub(r"[^a-z0-9]", "", td_lower)
    if not td_clean:
        return None

    best_idx = None
    best_score = 0.0
    for col_idx, hdr in uzio_col_headers.items():
        if col_idx < UZIO_FIRST_DATA_COL:
            continue
        hdr_lower = str(hdr).lower()
        hdr_clean = re.sub(r"[^a-z0-9]", "", hdr_lower)
        if not hdr_clean:
            continue
        score = difflib.SequenceMatcher(None, td_clean, hdr_clean).ratio()
        td_words = set(re.findall(r"[a-z0-9]+", td_lower))
        hdr_words = set(re.findall(r"[a-z0-9]+", hdr_lower))
        overlap = td_words & hdr_words
        if overlap:
            score += 0.15 * len(overlap)
        if "medicare" in td_words and "medicare" in hdr_words: score += 0.3
        if ("soc" in td_words or "ss" in td_words) and "social" in hdr_words: score += 0.3
        if ("fit" in td_words or "fed" in td_words) and "federal" in hdr_words: score += 0.3
        if "401k" in td_lower and "401k" in hdr_lower: score += 0.3
        if "regular" in td_words and "regular" in hdr_words: score += 0.3
        if "overtime" in td_words and "overtime" in hdr_words: score += 0.3
        if "bonus" in td_words and "bonus" in hdr_words: score += 0.3
        if "futa" in td_words and "futa" in hdr_words: score += 0.3
        if "sui" in td_words and "sui" in hdr_words: score += 0.3
        if score > best_score and score >= 0.7:
            best_score = score
            best_idx = col_idx
    return best_idx


def parse_filename_dates(filename):
    """Pull pay-period start/end and pay date from a Paycom filename like
    '...Pay Period 01112026 01172026 Pay Date 01232026.xlsx'.
    Returns (start, end, pay_date) as MM/DD/YYYY strings or (None, None, None).
    """
    name = (filename or "").rsplit(".", 1)[0]
    m = re.search(r"Pay Period\s+(\d{8})\s+(\d{8})\s*Pay Date\s+(\d{8})", name, re.IGNORECASE)
    if m:
        fmt = lambda d: f"{d[:2]}/{d[2:4]}/{d[4:]}"
        return fmt(m.group(1)), fmt(m.group(2)), fmt(m.group(3))
    return None, None, None


def reformat_name(raw_name):
    if not raw_name or not isinstance(raw_name, str):
        return raw_name or ""
    parts = re.split(r"\s{2,}", raw_name.strip())
    if len(parts) == 2:
        return f"{parts[0]}, {parts[1]}"
    return raw_name.strip()


def read_uzio_template(content):
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[wb.sheetnames[-1]]
    for sname in wb.sheetnames:
        if "payroll" in sname.lower() and "instruction" not in sname.lower():
            ws = wb[sname]
            break
    section_headers = {
        cell.column: str(cell.value).strip()
        for cell in ws[UZIO_SECTION_ROW] if cell.value
    }
    column_headers = {
        cell.column: str(cell.value).strip()
        for cell in ws[UZIO_HEADER_ROW] if cell.value
    }
    return section_headers, column_headers, wb, ws


def read_paycom_files(paycom_files_data):
    """paycom_files_data: list of (content_bytes, filename).
    Returns (paycom_data list, all_type_combos set of (tc, td, cd)).
    """
    paycom_data = []
    all_type_combos = set()
    for content, filename in paycom_files_data:
        pp_start, pp_end, pay_date = parse_filename_dates(filename)
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            rd = dict(zip(headers, row))
            rows.append(rd)
            tc = str(rd.get("Type Code", "") or "").strip()
            td = str(rd.get("Type Description", "") or "").strip()
            cd = str(rd.get("Code Description", "") or "").strip()
            if tc:
                all_type_combos.add((tc, td, cd))
        paycom_data.append({
            "filename": filename,
            "pp_start": pp_start,
            "pp_end": pp_end,
            "pay_date": pay_date,
            "rows": rows,
        })
    return paycom_data, all_type_combos


def generate_output(paycom_data, mapping, net_pay_col_idx):
    """Aggregate per (employee, file/pay-period) into one Uzio output row.

    `mapping` is keyed by (type_code, type_description) -> uzio_col_idx.
    """
    output_rows = []
    skipped_items = []
    validation_results = []

    for pf in paycom_data:
        ee_groups = defaultdict(list)
        for row in pf["rows"]:
            ee_code = str(row.get("EE Code", "") or "").strip()
            if ee_code:
                ee_groups[ee_code].append(row)

        for ee_code, rows in sorted(ee_groups.items()):
            out_row = {
                UZIO_EMPLOYEE_ID_COL: ee_code,
                UZIO_FULL_NAME_COL: reformat_name(str(rows[0].get("EE Name", "") or "")),
                UZIO_PP_START_COL: pf["pp_start"],
                UZIO_PP_END_COL: pf["pp_end"],
                UZIO_PAYCHECK_DATE_COL: pf["pay_date"],
            }

            net_pay_total = 0.0
            gross_earnings = 0.0
            total_ee_taxes = 0.0
            total_deductions = 0.0
            total_er_taxes = 0.0

            for row in rows:
                tc = str(row.get("Type Code", "") or "").strip()
                td = str(row.get("Type Description", "") or "").strip()
                cd = str(row.get("Code Description", "") or "").strip()
                try:
                    amt = float(row.get("Amount", 0) or 0)
                except (ValueError, TypeError):
                    amt = 0.0
                if cd == "Net Pay Distribution":
                    net_pay_total += amt
                    continue
                if cd == "Employee Benefits":
                    continue
                target_col = mapping.get((tc, td))
                if target_col is None:
                    if amt != 0:
                        skipped_items.append({
                            "Employee ID": ee_code,
                            "Pay Period Start": pf["pp_start"],
                            "Type Code": tc,
                            "Type Description": td,
                            "Code Description": cd,
                            "Amount": round(amt, 2),
                        })
                    continue
                out_row[target_col] = out_row.get(target_col, 0) + amt
                section = PAYCOM_CATEGORY_TO_SECTION.get(cd, "")
                if section == "Earnings":
                    gross_earnings += amt
                elif section == "Employee Taxes":
                    total_ee_taxes += amt
                elif section == "Employer Taxes":
                    total_er_taxes += amt
                elif section == "Deductions":
                    total_deductions += amt

            if net_pay_col_idx:
                out_row[net_pay_col_idx] = net_pay_total

            expected_net = gross_earnings - total_ee_taxes - total_deductions
            if abs(expected_net - net_pay_total) > 0.02:
                validation_results.append({
                    "Employee ID": ee_code,
                    "Pay Period": f"{pf['pp_start']} - {pf['pp_end']}",
                    "Gross Earnings": round(gross_earnings, 2),
                    "Employee Taxes": round(total_ee_taxes, 2),
                    "Deductions": round(total_deductions, 2),
                    "Expected Net": round(expected_net, 2),
                    "Actual Net Pay": round(net_pay_total, 2),
                    "Difference": round(expected_net - net_pay_total, 2),
                })

            output_rows.append(out_row)

    output_rows.sort(key=lambda r: (str(r.get(UZIO_EMPLOYEE_ID_COL, "")), str(r.get(UZIO_PP_START_COL, ""))))
    return output_rows, skipped_items, validation_results


def write_output_excel(uzio_wb, uzio_ws, output_rows, uzio_col_headers):
    if uzio_ws.max_row >= UZIO_DATA_START_ROW:
        uzio_ws.delete_rows(UZIO_DATA_START_ROW, uzio_ws.max_row - UZIO_DATA_START_ROW + 1)
    max_col = max(uzio_col_headers.keys()) if uzio_col_headers else 86
    for row_idx, out_row in enumerate(output_rows):
        excel_row = UZIO_DATA_START_ROW + row_idx
        for col_idx in range(1, max_col + 1):
            val = out_row.get(col_idx)
            if val is not None:
                uzio_ws.cell(row=excel_row, column=col_idx, value=val)
    buf = io.BytesIO()
    uzio_wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def run_paycom_prior_payroll_generator(uzio_template_bytes, paycom_files_data, override_mapping=None):
    """End-to-end pipeline. Returns (filled_xlsx_bytes, summary_dict).

    override_mapping accepts EITHER a dict keyed by "type_code|type_description"
    string (since JSON object keys must be strings) OR a list of
    [type_code, type_description, uzio_col_idx] triples. Negative idx force-skips.
    """
    section_headers, uzio_col_headers, uzio_wb, uzio_ws = read_uzio_template(uzio_template_bytes)
    paycom_data, all_type_combos = read_paycom_files(paycom_files_data)
    if not any(pf["rows"] for pf in paycom_data):
        raise ValueError("No valid data rows found in Paycom source files")

    mapping = {}
    for tc, td, cd in all_type_combos:
        section = PAYCOM_CATEGORY_TO_SECTION.get(cd, "Deductions")
        if section.startswith("_"):
            continue
        guess = auto_guess_mapping(td, cd, uzio_col_headers)
        if guess is not None:
            mapping[(tc, td)] = guess

    def _apply_override(tc, td, idx):
        try:
            idx_i = int(idx)
        except Exception:
            return
        if idx_i >= UZIO_FIRST_DATA_COL:
            mapping[(tc, td)] = idx_i
        else:
            mapping.pop((tc, td), None)

    if isinstance(override_mapping, dict):
        for k, v in override_mapping.items():
            tc, _, td = str(k).partition("|")
            _apply_override(tc.strip(), td.strip(), v)
    elif isinstance(override_mapping, list):
        for item in override_mapping:
            if isinstance(item, (list, tuple)) and len(item) >= 3:
                _apply_override(str(item[0]).strip(), str(item[1]).strip(), item[2])

    net_pay_col_idx = None
    for col_idx, hdr in uzio_col_headers.items():
        if "net pay" in str(hdr).lower():
            net_pay_col_idx = col_idx
            break

    output_rows, skipped_items, validation_results = generate_output(
        paycom_data, mapping, net_pay_col_idx,
    )
    xlsx_bytes = write_output_excel(uzio_wb, uzio_ws, output_rows, uzio_col_headers)

    summary = {
        "input_file_summaries": [
            {
                "Filename": pf["filename"],
                "Pay Period Start": pf["pp_start"],
                "Pay Period End": pf["pp_end"],
                "Pay Date": pf["pay_date"],
                "Records": len(pf["rows"]),
            }
            for pf in paycom_data
        ],
        "mapping": {f"{tc}|{td}": idx for (tc, td), idx in mapping.items()},
        "mapping_count": len(mapping),
        "skipped_with_values": skipped_items[:200],
        "skipped_count": len(skipped_items),
        "validation_issues": validation_results[:200],
        "validation_issue_count": len(validation_results),
        "output_rows": len(output_rows),
        "unique_employees": len({r.get(UZIO_EMPLOYEE_ID_COL) for r in output_rows}),
        "pay_periods": len({(r.get(UZIO_PP_START_COL), r.get(UZIO_PP_END_COL)) for r in output_rows}),
        "net_pay_target_col": net_pay_col_idx,
    }
    return xlsx_bytes, summary
