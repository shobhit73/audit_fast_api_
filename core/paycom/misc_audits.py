import pandas as pd
import io
import re
from utils.audit_utils import norm_blank, try_parse_date, norm_id, normalize_space_and_case

# --- Paycom Withholding Audit ---
def _pivot_uzio_long_to_wide(df_long):
    id_col = next((c for c in df_long.columns if "employee_id" == c.lower() or "employee id" == c.lower()), df_long.columns[0])
    key_col = next((c for c in df_long.columns if "field_key" in c.lower()), "withholding_field_key")
    val_col = next((c for c in df_long.columns if "field_value" in c.lower()), "withholding_field_value")
    wide = df_long.pivot_table(index=id_col, columns=key_col, values=val_col, aggfunc="first").reset_index()
    wide = wide.rename(columns={id_col: "employee_id"})
    return wide

def run_paycom_withholding_audit(uzio_content, paycom_content, mapping_content):
    """
    Production-grade Paycom withholding audit.
    Returns 3 sheets: Summary, Field_Summary_By_Status, Comparison_Detail_AllFields
    """
    uzio_long = pd.read_csv(io.BytesIO(uzio_content), dtype=str).fillna("")
    uzio_wide = _pivot_uzio_long_to_wide(uzio_long)
    
    try: paycom = pd.read_csv(io.BytesIO(paycom_content), dtype=str).fillna("")
    except: paycom = pd.read_excel(io.BytesIO(paycom_content), dtype=str).fillna("")
    
    mapping = pd.read_excel(io.BytesIO(mapping_content), dtype=str) if mapping_content else pd.DataFrame(columns=["Uzio Field Key", "PayCom Column"])

    p_id_col = next((c for c in paycom.columns if "Employee_Code" in c or "EE Code" in c.upper()), paycom.columns[0])
    paycom[p_id_col] = paycom[p_id_col].apply(norm_id)
    uzio_wide["employee_id"] = uzio_wide["employee_id"].apply(norm_id)

    paycom_ids = set(paycom[p_id_col].dropna())
    uzio_ids = set(uzio_wide["employee_id"].dropna())

    # Detect status column for Active/Terminated split
    paycom_status_col = next((c for c in paycom.columns if "status" in c.lower() and "employment" in c.lower()), None)
    
    merged = pd.merge(paycom, uzio_wide, left_on=p_id_col, right_on="employee_id", how="outer")
    
    rows = []
    for _, row in merged.iterrows():
        eid = row.get(p_id_col) or row.get("employee_id")
        if not eid or str(eid).strip() == "": continue
        
        emp_status = "ACTIVE"
        if paycom_status_col and paycom_status_col in merged.columns:
            s = str(row.get(paycom_status_col, "")).lower()
            if any(x in s for x in ["term", "inactive", "retired"]): emp_status = "TERMINATED"

        for _, m_row in mapping.iterrows():
            uz_key = m_row.get("Uzio Field Key")
            pc_col = m_row.get("PayCom Column")
            if not uz_key or not pc_col: continue
            if uz_key not in merged.columns and pc_col not in merged.columns: continue

            u_v = str(row.get(uz_key, "")).strip() if uz_key in merged.columns else ""
            p_v = str(row.get(pc_col, "")).strip() if pc_col in merged.columns else ""
            match = normalize_space_and_case(u_v) == normalize_space_and_case(p_v)

            status = "Data Match" if match else ("Value missing in Uzio (Paycom has value)" if not u_v and p_v else
                     ("Value missing in Paycom (Uzio has value)" if u_v and not p_v else "Data Mismatch"))
            rows.append({
                "Employee ID": eid, "Employment Status": emp_status,
                "Field": uz_key, "Paycom Column": pc_col,
                "UZIO_Value": u_v, "PAYCOM_Value": p_v, "Status": status
            })

    comparison_detail = rows
    df_cd = pd.DataFrame(rows)

    # Field_Summary_By_Status
    field_summary = []
    if not df_cd.empty:
        pivot = df_cd.pivot_table(index="Field", columns="Status", values="Employee ID", aggfunc="count", fill_value=0)
        for c in ["Data Match", "Data Mismatch", "Value missing in Uzio (Paycom has value)", "Value missing in Paycom (Uzio has value)"]:
            if c not in pivot.columns: pivot[c] = 0
        pivot["Total"] = pivot.sum(axis=1)
        field_summary = pivot.reset_index().to_dict(orient="records")

    total_mismatch_rows = sum(1 for r in rows if r["Status"] != "Data Match")
    active_mismatch_rows = sum(1 for r in rows if r["Status"] != "Data Match" and r["Employment Status"] == "ACTIVE")
    
    # Missing employees
    missing_in_uzio = paycom[~paycom[p_id_col].isin(uzio_ids)].copy()

    summary = [
        {"Metric": "Total UZIO Employees", "Value": len(uzio_ids)},
        {"Metric": "Total PAYCOM Employees", "Value": len(paycom_ids)},
        {"Metric": "Employees in both", "Value": len(uzio_ids & paycom_ids)},
        {"Metric": "Employees only in UZIO", "Value": len(uzio_ids - paycom_ids)},
        {"Metric": "Employees only in PAYCOM", "Value": len(paycom_ids - uzio_ids)},
        {"Metric": "Fields Compared", "Value": len(mapping)},
        {"Metric": "Total Comparisons (field-level rows)", "Value": len(rows)},
        {"Metric": "Total mismatches (mapped only)", "Value": total_mismatch_rows},
        {"Metric": "Active mismatches (mapped only)", "Value": active_mismatch_rows},
        {"Metric": "Missing in UZIO (Paycom employees)", "Value": len(missing_in_uzio)},
    ]

    return {
        "Summary": summary,
        "Field_Summary_By_Status": field_summary,
        "Comparison_Detail_AllFields": comparison_detail,
    }


# --- Paycom Emergency Audit ---
def _norm_str_ec(x):
    return str(x).strip() if x is not None and not (isinstance(x, float) and pd.isna(x)) else ""

def _norm_phone_ec(x):
    if not x or (isinstance(x, float) and pd.isna(x)): return ""
    digits = re.sub(r"\D", "", str(x))
    if len(digits) == 11 and digits.startswith("1"): digits = digits[1:]
    return digits

def _norm_relation_ec(x):
    return _norm_str_ec(x).upper()

def _compare_contact_field(field, u_val, p_val):
    u_s = str(u_val).strip().lower()
    p_s = str(p_val).strip().lower()
    if u_s == p_s: return True
    if field == "Phone":
        u_p = _norm_phone_ec(u_val); p_p = _norm_phone_ec(p_val)
        if u_p == p_p: return True
        if u_p and p_p and (u_p in p_p or p_p in u_p): return True
    if field == "Relationship":
        if u_s in ["spouse","husband","wife"] and p_s in ["spouse","husband","wife"]: return True
        if u_s in ["mother","father","parent"] and p_s in ["mother","father","parent"]: return True
        if "child" in u_s and "child" in p_s: return True
    return False

def run_paycom_emergency_audit(uzio_content, paycom_content):
    """
    Production-grade Paycom emergency contact audit.
    Returns 2 sheets: Emergency_Contact_Audit, Summary
    """
    df_uzio = pd.read_excel(io.BytesIO(uzio_content), header=1, dtype=str)
    try: df_paycom = pd.read_csv(io.BytesIO(paycom_content), dtype=str)
    except: df_paycom = pd.read_excel(io.BytesIO(paycom_content), dtype=str)
    df_paycom.columns = [str(c).strip().replace("\n", " ") for c in df_paycom.columns]

    u_eid = next((c for c in df_uzio.columns if "Employee ID" in c), "Employee ID")
    u_name = next((c for c in df_uzio.columns if "Name" in c and "Company" not in c), "Name")
    u_rel = next((c for c in df_uzio.columns if "Relationship" in c), "Relationship")
    u_phone = next((c for c in df_uzio.columns if "Phone" in c), "Phone")

    empid_col = next((c for c in df_paycom.columns if "Employee_Code" in c or "Employee ID" in c), df_paycom.columns[0])
    p_maps = []
    for i in range(1, 4):
        p_maps.append({
            "Name": next((c for c in df_paycom.columns if f"Emergency_{i}_Contact" in c), None),
            "Relation": next((c for c in df_paycom.columns if f"Emergency_{i}_Relationship" in c), None),
            "Phone": next((c for c in df_paycom.columns if f"Emergency_{i}_Phone" in c), None),
            "Language": next((c for c in df_paycom.columns if f"Emergency_{i}_Language" in c), None),
        })

    uzio_data, uzio_all_ids = {}, set()
    for _, row in df_uzio.iterrows():
        eid = norm_id(row.get(u_eid, ""))
        if not eid: continue
        uzio_all_ids.add(eid)
        contact = {"Name": _norm_str_ec(row.get(u_name)), "Relation": _norm_relation_ec(row.get(u_rel)),
                   "Phone": _norm_phone_ec(row.get(u_phone)), "RawPhone": _norm_str_ec(row.get(u_phone)), "Language": ""}
        if contact["Name"] or contact["Phone"]:
            uzio_data.setdefault(eid, []).append(contact)

    paycom_data, paycom_all_ids = {}, set()
    for _, row in df_paycom.iterrows():
        eid = norm_id(row.get(empid_col, ""))
        if not eid: continue
        paycom_all_ids.add(eid)
        for pm in p_maps:
            if not pm["Name"] and not pm["Phone"]: continue
            contact = {
                "Name": _norm_str_ec(row.get(pm["Name"])) if pm["Name"] else "",
                "Relation": _norm_relation_ec(row.get(pm["Relation"])) if pm["Relation"] else "",
                "Phone": _norm_phone_ec(row.get(pm["Phone"])) if pm["Phone"] else "",
                "RawPhone": _norm_str_ec(row.get(pm["Phone"])) if pm["Phone"] else "",
                "Language": _norm_str_ec(row.get(pm["Language"])) if pm["Language"] else "",
            }
            if contact["Name"] or contact["Phone"]:
                paycom_data.setdefault(eid, []).append(contact)

    FIELDS = ["Name", "Relationship", "Phone"]
    rows = []
    for eid in sorted(set(uzio_data.keys()) | set(paycom_data.keys())):
        u_contacts = uzio_data.get(eid, [])
        p_contacts = paycom_data.get(eid, [])

        if not u_contacts and p_contacts:
            status = "Employee ID not in Uzio (present in paycom)" if eid not in uzio_all_ids else "Missing in Uzio"
            for p in p_contacts:
                for f in FIELDS:
                    rows.append({"Employee ID": eid, "Status": status, "Field": f, "Uzio Value": "", "Paycom Value": p.get(f, p.get("RawPhone") if f=="Phone" else "")})
            continue
        if u_contacts and not p_contacts:
            status = "Employee ID not in Paycom (Present in uzio)" if eid not in paycom_all_ids else "Missing in Paycom"
            for u in u_contacts:
                for f in FIELDS:
                    rows.append({"Employee ID": eid, "Status": status, "Field": f, "Uzio Value": u.get("RawPhone") if f=="Phone" else u.get(f,""), "Paycom Value": ""})
            continue

        u_pending, p_pending, matched_pairs = u_contacts[:], p_contacts[:], []
        for u in list(u_pending):
            match = next((p for p in p_pending if u["Name"].lower() == p["Name"].lower()), None)
            if match: matched_pairs.append((u, match)); u_pending.remove(u); p_pending.remove(match)
        for u in list(u_pending):
            if not u["Phone"]: continue
            match = next((p for p in p_pending if p["Phone"] and u["Phone"] == p["Phone"]), None)
            if match: matched_pairs.append((u, match)); u_pending.remove(u); p_pending.remove(match)

        for u, p in matched_pairs:
            for f in FIELDS:
                u_val = u.get("RawPhone") if f == "Phone" else u.get(f, "")
                p_val = p.get("RawPhone") if f == "Phone" else p.get(f, "")
                match_r = _compare_contact_field(f, u.get(f,""), p.get(f,""))
                rows.append({"Employee ID": eid, "Status": "Data Match" if match_r else "Data Mismatch", "Field": f, "Uzio Value": u_val, "Paycom Value": p_val})
            rows.append({"Employee ID": eid, "Status": "Info Only", "Field": "Language", "Uzio Value": "N/A", "Paycom Value": p.get("Language","")})
        for u in u_pending:
            for f in FIELDS:
                rows.append({"Employee ID": eid, "Status": "Missing in Paycom", "Field": f, "Uzio Value": u.get("RawPhone") if f=="Phone" else u.get(f,""), "Paycom Value": ""})
        for p in p_pending:
            for f in FIELDS:
                rows.append({"Employee ID": eid, "Status": "Missing in Uzio", "Field": f, "Uzio Value": "", "Paycom Value": p.get("RawPhone") if f=="Phone" else p.get(f,"")})

    df_res = pd.DataFrame(rows)
    summary = []
    if not df_res.empty:
        summ = df_res.groupby(["Status", "Field"]).size().reset_index(name="Count")
        summary = summ.to_dict(orient="records")

    return {"Emergency_Contact_Audit": rows, "Summary": summary}


# --- Paycom Timeoff Audit ---
def run_paycom_timeoff_audit(uzio_content, paycom_content):
    """
    Production-grade Paycom timeoff audit.
    Reads ADP balance export and updates Uzio template.
    """
    try:
        import openpyxl
        paycom_df = pd.read_excel(io.BytesIO(paycom_content), dtype=str)
        p_id_col = next((c for c in paycom_df.columns if "employee" in c.lower() and ("code" in c.lower() or "id" in c.lower())), paycom_df.columns[0])
        p_bal_col = next((c for c in paycom_df.columns if "balance" in c.lower()), None)
        if not p_bal_col:
            return {"message": "Could not find Balance column in Paycom file", "status": "Error"}

        paycom_df["_clean_id"] = paycom_df[p_id_col].apply(lambda x: str(x).strip().lstrip("0").replace(".0",""))
        balance_map = {}
        for _, row in paycom_df.iterrows():
            eid = row["_clean_id"]
            val = row.get(p_bal_col)
            try: balance_map[eid] = float(str(val).replace(",","").strip())
            except: pass

        wb = openpyxl.load_workbook(io.BytesIO(uzio_content))
        ws = wb.worksheets[1]  # Time Off Details
        header_row = 4
        idx_id, idx_bal = None, None
        for cell in ws[header_row]:
            v = str(cell.value).strip() if cell.value else ""
            if "Employee ID" in v: idx_id = cell.column
            elif "Opening Balance" in v: idx_bal = cell.column

        updates = 0
        if idx_id and idx_bal:
            for r in range(header_row + 1, ws.max_row + 1):
                eid = str(ws.cell(row=r, column=idx_id).value or "").strip().lstrip("0").replace(".0","")
                if eid in balance_map:
                    ws.cell(row=r, column=idx_bal).value = balance_map[eid]
                    updates += 1

        return {"message": f"Updated {updates} balances in Uzio template", "status": "Success"}
    except Exception as e:
        return {"message": f"Timeoff audit error: {e}", "status": "Error"}
