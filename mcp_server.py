import json
import base64
import os
import numpy as np
import duckdb
from mcp.server.models import InitializationOptions
from mcp.server import NotificationOptions, Server
from mcp.server.sse import SseServerTransport
import mcp.types as types
import sys

# ── Drop-folder: user puts files here, Claude picks them up automatically ──
AUDIT_INBOX = r"C:\Users\shobhit.sharma\Desktop\Audit Files"

# --- Core Imports ---
from core.adp.total_comparison import run_adp_total_comparison
from core.adp.census_audit import run_adp_census_audit, ADP_FIELD_MAP
from core.adp.deduction_audit import run_adp_deduction_audit
from core.adp.payment_audit import run_adp_payment_audit
from core.adp.withholding_audit import run_adp_withholding_audit

# paycom_deduction_analyzer was deleted -- replaced by paycom_prior_payroll_setup_helper
from core.paycom.deduction_audit import run_paycom_deduction_audit
from core.paycom.total_comparison import run_paycom_total_comparison
from core.paycom.census_audit import run_paycom_census_audit, PAYCOM_FIELD_MAP
from core.paycom.withholding_audit import run_paycom_withholding_audit
from core.paycom.sql_master import run_paycom_sql_master
from core.paycom.payment_audit import run_paycom_payment_audit
from core.paycom.misc_audits import run_paycom_emergency_audit, run_paycom_timeoff_audit

from core.census.sanity_check import run_census_sanity_check, generate_corrected_census_xlsx
from core.adp.misc_audits import (
    run_adp_emergency_audit, run_adp_license_audit, run_adp_timeoff_audit
)
from core.adp.census_generator import run_adp_census_generation
from core.paycom.census_generator import run_paycom_census_generation
from core.adp.prior_payroll_sanity import run_adp_prior_payroll_sanity
from core.adp.prior_payroll_generator import run_adp_prior_payroll_generator
from core.paycom.prior_payroll_generator import run_paycom_prior_payroll_generator
from core.adp.selective_census_sync import run_adp_selective_census_sync, discover_mappings as adp_selective_discover
from core.paycom.selective_census_sync import run_paycom_selective_census_sync, discover_mappings as paycom_selective_discover
from core.common.paycom_consolidated_audit import run_paycom_consolidated_audit
from core.adp.prior_payroll_setup_helper import run_adp_prior_payroll_setup_helper, build_simplified_xlsx_bytes as _setup_helper_xlsx
from core.paycom.prior_payroll_setup_helper import run_paycom_prior_payroll_setup_helper
from utils.file_shape_guards import require_vendor

from starlette.applications import Starlette
from starlette.routing import Mount, Route

server = Server("audit-tool-server")

# ── Helpers ──────────────────────────────────────────────────────────────────

# ── Caching System for SQL Tools ─────────────────────────────────────────────
_DF_CACHE = {} # Key: (path, mtime, size, sheet_name) | Value: pd.DataFrame

def _get_cached_df(path, sheet_name=None):
    """Loads a file into a DataFrame with in-memory caching."""
    import os
    import pandas as pd
    from utils.audit_utils import find_header_and_data
    
    if not os.path.isfile(path):
        return None, f"Error: File '{path}' not found."
    
    try:
        stat = os.stat(path)
        cache_key = (path, stat.st_mtime, stat.st_size, sheet_name)
        
        if cache_key in _DF_CACHE:
            return _DF_CACHE[cache_key], None
            
        # Load the data
        if sheet_name:
            df = pd.read_excel(path, sheet_name=sheet_name) if path.lower().endswith(('.xlsx', '.xls')) else pd.read_csv(path)
        else:
            with open(path, "rb") as f:
                df, _, _ = find_header_and_data(f.read(), os.path.basename(path))
        
        # Limit cache size (keep last 5 files)
        if len(_DF_CACHE) > 5:
            _DF_CACHE.clear()
            
        _DF_CACHE[cache_key] = df
        return df, None
    except Exception as e:
        return None, str(e)

def _json_default(o):
    """Custom JSON encoder for numpy types and other non-serializable objects."""
    if isinstance(o, (np.integer,)):
        return int(o)
    if isinstance(o, (np.floating,)):
        return float(o)
    if isinstance(o, np.ndarray):
        return o.tolist()
    if hasattr(o, "isoformat"):  # handles datetime, Timestamp, etc.
        return o.isoformat()
    # Handle pandas types specifically if numpy check isn't enough
    if hasattr(o, "item") and callable(o.item):
        return o.item()
    raise TypeError(f"Object of type {type(o).__name__} is not JSON serializable")

def save_results_to_excel(results, name_prefix):
    """Saves results (list of dicts OR dict of lists) to an Excel file on the Desktop."""
    import pandas as pd
    from datetime import datetime
    
    if not results:
        return {"summary": "No results", "file_path": None}
        
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"{name_prefix}_{stamp}.xlsx"
    
    # Ensure the Audit Files directory exists
    if not os.path.exists(AUDIT_INBOX):
        os.makedirs(AUDIT_INBOX, exist_ok=True)
        
    out_path = os.path.join(AUDIT_INBOX, filename)
    
    if isinstance(results, list):
        df = pd.DataFrame(results)
        df.to_excel(out_path, index=False)
        summary = {
            "total_rows": len(results),
            "file_path": out_path,
            "message": f"Full report saved to 'Audit Files' folder as {filename}.",
            "data": results if len(results) < 2000 else results[:500],
            "note": "Full data returned in response." if len(results) < 2000 else "Data truncated due to size."
        }
        if "Status" in df.columns:
            summary["counts_by_status"] = df["Status"].value_counts().to_dict()
        return summary
        
    elif isinstance(results, dict):
        # Handle dict of lists (multiple sheets)
        with pd.ExcelWriter(out_path) as writer:
            summary_info = {"file_path": out_path, "message": f"Report saved to 'Audit Files' folder as {filename}."}
            for sheet_name, data in results.items():
                if isinstance(data, list) and data:
                    df = pd.DataFrame(data)
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
                    summary_info[f"{sheet_name}_count"] = len(data)
                    # Return full data to Claude if it's reasonably sized (< 2000 rows)
                    if len(data) < 2000:
                        summary_info[sheet_name] = data
                    else:
                        summary_info[sheet_name] = data[:500]
                        summary_info[f"{sheet_name}_note"] = "Data truncated in response due to size. Full data in Excel file."
                else:
                    summary_info[sheet_name] = data
            return summary_info
            
    return {"error": "Unsupported results format"}

def load_file(arguments: dict, path_key: str, b64_key: str) -> bytes:
    """Load a file from a local path OR fall back to base64-encoded content."""
    path = arguments.get(path_key)
    if path:
        path = path.strip().strip('"')
        with open(path, "rb") as f:
            return f.read()
    b64 = arguments.get(b64_key)
    if b64:
        return base64.b64decode(b64)
    return b""

def load_files_list(arguments: dict, paths_key: str, b64_key: str):
    """Load a list of files from local paths OR base64 list. Returns list of (bytes, name)."""
    paths = arguments.get(paths_key, [])
    if paths:
        result = []
        for p in paths:
            p = p.strip().strip('"')
            with open(p, "rb") as f:
                result.append((f.read(), os.path.basename(p)))
        return result
    b64_list = arguments.get(b64_key, [])
    return [(base64.b64decode(b), f"file_{i}.xlsx") for i, b in enumerate(b64_list)]

def load_mappings_from_paths(paths):
    """Loads and merges mappings from a list of local file paths (CSV or Excel)."""
    import pandas as pd
    mappings = []
    for p in paths:
        p = p.strip().strip('"')
        if not os.path.isfile(p): continue
        try:
            # Determine category from filename
            cat = "Earnings"
            fname = os.path.basename(p).lower()
            if "deduction" in fname: cat = "Deductions"
            elif "contribution" in fname: cat = "Contributions"
            elif "tax" in fname: cat = "Taxes"
            
            df = pd.read_csv(p) if p.lower().endswith('.csv') else pd.read_excel(p)
            
            # Find Source and Uzio columns
            s_col = next((c for c in df.columns if "source" in str(c).lower() and "name" in str(c).lower()), None)
            u_col = next((c for c in df.columns if "uzio" in str(c).lower() and ("name" in str(c).lower() or "description" in str(c).lower())), None)
            
            if s_col and u_col:
                for _, row in df.iterrows():
                    mappings.append({
                        "Category": cat,
                        "ADP_Name": str(row[s_col]).strip(),
                        "UZIO_Name": str(row[u_col]).strip()
                    })
        except Exception as e:
            print(f"Error loading mapping {p}: {e}")
    return mappings

def copy_file_to_inbox(source_path):
    """Safely copies a file from anywhere on the system to the Audit Files inbox."""
    import shutil
    source_path = source_path.strip().strip('"')
    if not os.path.isfile(source_path):
        return {"error": f"Source file '{source_path}' not found."}
    
    # Ensure inbox exists
    if not os.path.exists(AUDIT_INBOX):
        os.makedirs(AUDIT_INBOX, exist_ok=True)
        
    dest_path = os.path.join(AUDIT_INBOX, os.path.basename(source_path))
    try:
        shutil.copy2(source_path, dest_path)
        return {
            "success": True,
            "message": f"File successfully copied to inbox.",
            "source": source_path,
            "destination": dest_path
        }
    except Exception as e:
        return {"error": f"Failed to copy file: {str(e)}"}

def apply_data_corrections(file_path, corrections_list):
    """
    Surgically updates specific cells in an Excel file using openpyxl to preserve formatting.
    corrections_list: list of dicts like {'id': '123', 'column': 'Status', 'value': 'Inactive'}
    """
    import openpyxl
    from utils.audit_utils import norm_id, norm_colname
    
    file_path = file_path.strip().strip('"')
    if not os.path.isfile(file_path):
        return {"error": f"File '{file_path}' not found."}
    
    try:
        wb = openpyxl.load_workbook(file_path)
        # Prefer the sanity-output data sheet when present; fall back to whatever
        # was active when the file was last saved.
        if "Corrected Census" in wb.sheetnames:
            ws = wb["Corrected Census"]
        else:
            ws = wb.active

        # 1. Identify header row (some files have junk rows at the top)
        header_row_idx = 1
        id_keywords = ["employee id", "employee code", "associate id", "file #", "id#"]
        
        for r in range(1, 20): # Peek first 20 rows
            row_vals = [norm_colname(cell.value).lower() for cell in ws[r]]
            if any(any(k in v for k in id_keywords) for v in row_vals):
                header_row_idx = r
                norm_headers = row_vals
                headers = [str(cell.value) for cell in ws[r]]
                break
        else:
            # Fallback to row 1 if no keywords found
            headers = [str(cell.value) for cell in ws[1]]
            norm_headers = [norm_colname(h).lower() for h in headers]
        
        # Find ID column index
        id_col_indices = [i for i, h in enumerate(norm_headers) if any(k in h for k in id_keywords)]
        if not id_col_indices:
            return {"error": f"Could not identify Employee ID column. Found headers in row {header_row_idx}: {headers}"}
        id_col_idx = id_col_indices[0]
        
        results = []
        for corr in corrections_list:
            target_id = norm_id(corr.get('id'))
            target_col = norm_colname(corr.get('column')).lower()
            new_val = corr.get('value')
            
            # Find target column - error on ambiguity rather than silently
            # picking the first match (e.g. "FLSA" matches both "FLSA Description"
            # and "FLSA Code").
            matches = [(i, headers[i]) for i, h in enumerate(norm_headers) if target_col in h]
            if not matches:
                results.append({"id": target_id, "status": "Error", "message": f"Column '{corr.get('column')}' not found."})
                continue
            if len(matches) > 1:
                cand_names = [m[1] for m in matches]
                results.append({
                    "id": target_id,
                    "status": "Error",
                    "message": f"Column '{corr.get('column')}' is ambiguous; matches multiple headers: {cand_names}. Use the full header name."
                })
                continue
            col_idx = matches[0][0]
            
            # 2. Find row and update
            found = False
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                cell_val = norm_id(ws.cell(row=row_idx, column=id_col_idx + 1).value)
                if cell_val == target_id:
                    ws.cell(row=row_idx, column=col_idx + 1).value = new_val
                    results.append({"id": target_id, "column": corr.get('column'), "status": "Success"})
                    found = True
                    break
            
            if not found:
                results.append({"id": target_id, "status": "Error", "message": "Employee ID not found in file."})
        
        # 3. Save as new file with suffix
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        base, ext = os.path.splitext(file_path)
        out_path = f"{base}_OVERRIDDEN_{stamp}{ext}"
        wb.save(out_path)
        
        return {
            "success": True,
            "output_file": out_path,
            "applied_changes": results
        }
        
    except Exception as e:
        import traceback
        return {"error": f"Correction failed: {str(e)}\n{traceback.format_exc()}"}

# ── Tool Definitions ──────────────────────────────────────────────────────────

PATH_DESC = "Full local file path (e.g. C:\\Users\\...\\file.xlsx). Preferred over base64 for large files."

@server.list_tools()
async def handle_list_tools() -> list[types.Tool]:
    return [
        # --- UTILITY TOOLS ---
        types.Tool(
            name="list_audit_files",
            description=(
                "[VENDOR-AGNOSTIC] [INPUT: optional directory path; defaults to "
                "C:/Users/shobhit.sharma/Desktop/Audit Files]\n"
                "Lists all files in the audit drop-folder (or any user-specified directory) "
                "with full paths, sizes, and last-modified timestamps. ALWAYS call this first "
                "before running any audit to discover which files the user has available - "
                "do not guess at filenames."
            ),
            inputSchema={
                "type": "object", 
                "properties": {
                    "directory_path": {
                        "type": "string",
                        "description": "Optional: Full local path to a folder to scan (e.g., C:\\Users\\...\\Happy Delivery). Defaults to Desktop/Audit Files."
                    }
                }
            },
        ),
        types.Tool(
            name="copy_to_audit_inbox",
            description=(
                "[VENDOR-AGNOSTIC] [INPUT: source_path = any local file path]\n"
                "Copies a file from any local directory (e.g. Downloads, a client folder) "
                "into the 'Audit Files' inbox so audit tools can find it consistently. "
                "Optional but recommended - audit tools accept any local path, but consolidating "
                "in the inbox keeps inputs and outputs together."
            ),
            inputSchema={
                "type": "object", 
                "properties": {
                    "source_path": {
                        "type": "string",
                        "description": "Full local path to the source file (e.g., C:\\Users\\...\\Downloads\\report.xlsx)"
                    }
                },
                "required": ["source_path"]
            },
        ),
        types.Tool(
            name="apply_data_corrections",
            description=(
                "Performs surgical row-level overrides on a post-sanity census file, "
                "keyed off Employee ID (Paycom) or Associate ID (ADP). Preserves all "
                "original Excel formatting (colors, fonts, borders, column widths). "
                "Output: a NEW file '<base>_OVERRIDDEN_<timestamp>.xlsx' next to the "
                "input - the input is never modified in place.\n\n"
                "WORKFLOW (file_path resolution, in priority order):\n"
                "  1. If the prior turn ran 'adp_census_sanity' or "
                "'paycom_census_sanity', use the 'output_file' path returned in that "
                "response - that is the canonical post-sanity file.\n"
                "  2. Otherwise, call 'list_audit_files' and pick the most recent "
                "'<Vendor>_Cleaned_*.xlsx' (vendor = ADP or Paycom) by filename "
                "timestamp from the Audit Files inbox "
                "(C:\\Users\\<user>\\Desktop\\Audit Files).\n"
                "  3. Only ask the user if multiple vendors have recent cleaned files "
                "and the override list itself does not disambiguate.\n\n"
                "COLUMN-NAME RULE: Use the FULL header from the file (e.g. "
                "'FLSA Description', not 'FLSA'). Matching is substring-based, so "
                "ambiguous fragments will now error out listing all candidate headers "
                "rather than silently picking the first one."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {
                        "type": "string",
                        "description": (
                            "Full local path to the post-sanity Excel file. Typically "
                            "the 'output_file' returned by adp_census_sanity / "
                            "paycom_census_sanity, or the most recent "
                            "<Vendor>_Cleaned_*.xlsx in the Audit Files inbox."
                        ),
                    },
                    "corrections": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "id": {
                                    "type": "string",
                                    "description": (
                                        "Employee ID (Paycom) or Associate ID (ADP). "
                                        "Leading zeros and trailing '.0' are normalized "
                                        "automatically."
                                    ),
                                },
                                "column": {
                                    "type": "string",
                                    "description": (
                                        "Full column header as it appears in the file's "
                                        "header row (e.g. 'FLSA Description', "
                                        "'Primary Address: Zip / Postal Code'). Partial "
                                        "names like 'FLSA' will error if they match "
                                        "multiple columns."
                                    ),
                                },
                                "value": {
                                    "type": "string",
                                    "description": "New value to write into the cell.",
                                },
                            },
                            "required": ["id", "column", "value"],
                        },
                        "description": (
                            "List of row-level overrides to apply. Each item targets "
                            "one cell, identified by (id, column)."
                        ),
                    },
                },
                "required": ["file_path", "corrections"],
            },
        ),

        # --- ADP TOOLS ---
        types.Tool(
            name="adp_total_comparison",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [adp_file_paths: 1+ ADP Prior Payroll "
                "Register Reports] [uzio_file_path: UZIO Master / Custom Report]\n"
                "[DO NOT confuse the slots - ADP file goes in adp_file_paths, UZIO file goes "
                "in uzio_file_path. Wrong placement will produce nonsense matches.]\n\n"
                "Performs a complete payroll total comparison between ADP and Uzio reports, "
                "producing up to 7 sheets: Full Comparison, Mismatches Only, Employee Mismatches, "
                "All Employee Details, Duplicate Pay Periods, Pay Stub Counts, and Tax Rate Verification.\n\n"
                "[MAPPINGS ARE REQUIRED FOR FULL OUTPUT] Without mappings, only Pay Stub Counts "
                "and Tax Rate Verification will populate - the four comparison sheets (Full Comparison, "
                "Mismatches Only, Employee Mismatches, All Employee Details) will be silently empty "
                "and not written to the Excel file. The Streamlit version asks the user to upload "
                "4 mapping files (Earnings, Deductions, Contributions, Taxes) and you MUST do the "
                "equivalent here.\n\n"
                "[BEFORE CALLING THIS TOOL] If the user has not provided mapping files or "
                "mappings_json, STOP and ASK them: 'Do you have the 4 mapping files (Earnings, "
                "Deductions, Contributions, Taxes mapping files - the same ones the Streamlit "
                "tool asks for)? Please share their paths.' Do NOT proceed with an empty mappings "
                "list - the resulting report will be missing the comparison sheets and the user "
                "will think the tool is broken.\n\n"
                "HINT: Always call 'list_audit_files' first to identify the correct file paths."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "adp_file_paths": {"type": "array", "items": {"type": "string"}, "description": PATH_DESC},
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "mapping_file_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": (
                            "REQUIRED for full report. Local paths to the 4 mapping files: "
                            "Earnings (Source Earning Code Name -> Uzio Earning Code Name), "
                            "Deductions (Source Deduction Code Name -> Uzio Deduction Code Name), "
                            "Contributions (Source Contribution Code Name -> Uzio Contribution Code Name), "
                            "and Taxes (Source Tax Code Name -> Uzio Tax Code Description). "
                            "If the user has not provided these, ASK before calling - do NOT call "
                            "this tool with an empty list."
                        ),
                    },
                    "mappings_json": {
                        "type": "string",
                        "description": (
                            "Alternative to mapping_file_paths. Flat JSON array of objects with "
                            "keys Category ('Earnings' | 'Deductions' | 'Contributions' | 'Taxes'), "
                            "ADP_Name, UZIO_Name. Use only if the user provides mappings inline."
                        ),
                    },
                    "adp_files_base64": {"type": "array", "items": {"type": "string"}, "description": "Fallback: base64 encoded ADP files"},
                    "uzio_file_base64": {"type": "string", "description": "Fallback: base64 encoded Uzio file"},
                },
                "required": [],
            },
        ),
        types.Tool(
            name="adp_census_audit",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [uzio_file_path: UZIO Census Custom Report] "
                "[adp_file_path: ADP Census export (.xlsx/.csv)]\n"
                "[DO NOT swap the slots - mixing them produces nonsense.]\n"
                "Audits employee census data between Uzio and ADP to find mismatches in names, "
                "emails, addresses, hire/term dates, status, FLSA, etc."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string", "description": "Fallback: base64 Uzio file"},
                    "adp_raw_base64": {"type": "string", "description": "Fallback: base64 ADP file"},
                },
            },
        ),
        types.Tool(
            name="adp_deduction_audit",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [uzio_file_path: UZIO Deduction Report] "
                "[adp_file_path: ADP Deduction Register / Prior Payroll]\n"
                "[REQUIRED: mapping_json - JSON object mapping ADP deduction codes to UZIO codes]\n"
                "Compares deduction amounts between Uzio and ADP reports per employee, per code."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "mapping_json": {"type": "string", "description": "JSON mapping of deduction codes"},
                    "uzio_raw_base64": {"type": "string"},
                    "adp_raw_base64": {"type": "string"},
                },
                "required": ["mapping_json"],
            },
        ),
        types.Tool(
            name="adp_payment_audit",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [uzio_file_path: UZIO Payment Report] "
                "[adp_file_path: ADP Payment Register / Direct Deposit Report]\n"
                "Audits payment methods (direct deposit / check) and amounts between Uzio and ADP."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "adp_raw_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="adp_withholding_audit",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [uzio_file_path: UZIO Withholding Report] "
                "[adp_file_path: ADP Withholding / W-4 Report]\n"
                "Audits federal + state tax withholding settings (filing status, allowances, "
                "extra withholding, exemptions) between Uzio and ADP per employee."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "adp_raw_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="adp_census_sanity",
            description=(
                "[VENDOR: ADP only] [INPUT: ADP Census export (.xlsx/.csv) with columns "
                "Associate ID, Legal First Name, FLSA Description, etc.]\n"
                "[DO NOT USE FOR: UZIO Master / UZIO Census Template / Paycom Census - the "
                "runtime guard will refuse those. For Paycom census use 'paycom_census_sanity'.]\n\n"
                "Applies opt-in auto-corrections to an ADP Census export. "
                "MANDATORY: For stability, always use copy_to_audit_inbox first and then use 'file_path'. "
                "Do NOT use 'file_base64' for files > 1MB."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "file_base64": {"type": "string", "description": "Fallback: base64 encoded ADP Census export"},
                    "filename": {"type": "string"},
                    "fix_flsa": {"type": "boolean"},
                    "fix_emails": {"type": "boolean"},
                    "fix_job_title": {"type": "boolean"},
                    "fix_driver_smart": {"type": "boolean"},
                    "fix_license": {"type": "boolean"},
                    "fix_status": {"type": "boolean"},
                    "fix_type": {"type": "boolean"},
                    "fix_dol_status": {"type": "boolean"},
                    "fix_leave_to_active": {"type": "boolean"},
                    "fix_blank_jt_to_driver": {"type": "boolean"},
                    "fix_std_hours": {"type": "boolean"},
                    "rename_std_hours": {"type": "boolean"},
                    "fix_zip": {"type": "boolean"},
                    "rename_zip_col": {"type": "boolean"},
                    "replace_gender_col": {"type": "boolean"},
                    "sort_by_manager": {"type": "boolean"},
                },
            },
        ),
        types.Tool(
            name="adp_selective_census_sync",
            description=(
                "Updates ONLY the requested columns in a pre-filled Uzio Census Template "
                "(.xlsm) using a fresh ADP census export. Employees not present in the ADP "
                "source are left untouched; everything outside selected_uzio_cols is also "
                "preserved exactly as it was in the template (the .xlsm's VBA, instructions, "
                "and other sheets pass through unchanged).\n\n"
                "Selected columns must be keys of the Uzio raw mapping (e.g. 'Employee ID*', "
                "'Employee First Name*', 'Date of Hire*', 'Standard Hours*', 'Primary "
                "Address: Zip / Postal Code', etc.). Job Title and Work Location are special: "
                "if the caller wants those synced too, pass an explicit job_title_mapping / "
                "work_location_mapping dict (source_value -> Uzio_value). Pass {} to seed "
                "automatically from whatever mapping is already present in the template "
                "(extract_mappings_from_uzio walks the existing template to learn the "
                "convention). Pass null/omit to skip syncing those columns.\n\n"
                "fix_options carries the same toggle keys as adp_census_sanity (fix_emails, "
                "fix_license, fix_status, fix_type, fix_job_title, ...).\n\n"
                "Set discover_only=true to skip writing and instead return the seed "
                "mappings + unique source values so the caller can review before applying."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_base64": {"type": "string", "description": "Fallback: base64 ADP census file"},
                    "filename": {"type": "string", "description": "Optional filename hint when using base64."},
                    "uzio_template_path": {"type": "string", "description": PATH_DESC + " (the pre-filled .xlsm)"},
                    "uzio_template_base64": {"type": "string", "description": "Fallback: base64 Uzio template"},
                    "selected_uzio_cols": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Keys from UZIO_RAW_MAPPING -- the columns to overwrite (e.g. 'Employee SSN', 'Date of Hire*').",
                    },
                    "job_title_mapping": {
                        "type": "object",
                        "description": "Optional {source_job_title: uzio_job_title} dict. Pass {} to seed from template.",
                        "additionalProperties": {"type": "string"},
                    },
                    "work_location_mapping": {
                        "type": "object",
                        "description": "Optional {source_location: uzio_location} dict. Pass {} to seed from template.",
                        "additionalProperties": {"type": "string"},
                    },
                    "fix_options": {
                        "type": "object",
                        "description": "Optional auto-fix toggles (fix_emails, fix_license, fix_status, fix_type, fix_job_title, ...).",
                        "additionalProperties": {"type": "boolean"},
                    },
                    "discover_only": {
                        "type": "boolean",
                        "description": "If true, return only the seed mappings without writing the template.",
                    },
                },
                "required": ["selected_uzio_cols"],
            },
        ),
        types.Tool(
            name="adp_prior_payroll_generator",
            description=(
                "Generates a filled Uzio Prior Payroll Template (.xlsx) from 1-10 ADP "
                "Prior Payroll History files. Each ADP dynamic column is auto-mapped to a "
                "Uzio target column via a fuzzy-string heuristic (handles Medicare, Social "
                "Security, FIT, 401k, FUTA, SUI/SDI, regular/overtime/bonus, state income, "
                "etc.). The auto-mapping can be overridden per-column with override_mapping. "
                "Records are aggregated per (employee, pay-period-start), Net Pay is routed "
                "to the column whose Uzio header contains 'net pay', and a validation pass "
                "flags any employee-period where Gross - Taxes - Deductions != Net Pay.\n\n"
                "WORKFLOW: copy both the blank Uzio template and the ADP file(s) to the "
                "Audit Files inbox first, then pass the resulting paths. The output xlsx is "
                "written to the same inbox and its path returned in the response."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_template_path": {"type": "string", "description": PATH_DESC},
                    "uzio_template_base64": {"type": "string", "description": "Fallback: base64-encoded Uzio Prior Payroll Template (headers only)."},
                    "adp_file_paths": {
                        "type": "array", "items": {"type": "string"},
                        "description": "List of local paths to ADP Prior Payroll History .xlsx files (max 10).",
                    },
                    "adp_files_base64": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Fallback: base64-encoded ADP files (max 10).",
                    },
                    "override_mapping": {
                        "type": "object",
                        "description": (
                            "Optional {adp_column_name: uzio_column_index} override. Use a "
                            "negative integer to force-skip a column. Auto-guessed pairs are "
                            "kept for any ADP column not present in this object."
                        ),
                        "additionalProperties": {"type": "integer"},
                    },
                    "client_name": {
                        "type": "string",
                        "description": "Optional client name; used in the output filename.",
                    },
                },
            },
        ),
        types.Tool(
            name="adp_prior_payroll_sanity",
            description=(
                "[VENDOR: ADP only] [INPUT: ADP Prior Payroll Register Report (.xlsx/.csv)]\n"
                "[DO NOT USE FOR: UZIO Master / UZIO Custom Report / Paycom Prior Payroll - "
                "the runtime guard will refuse those with a clear error.]\n\n"
                "Cleans an ADP Prior Payroll export so it can be ingested by downstream APIs. "
                "Three independent fix-ups are applied as needed:\n"
                "  1. Drops interleaved 'Totals For Associate ID XYZ:' summary rows.\n"
                "  2. Detects and removes the bottom-of-file grand-total row where the last "
                "employee's ID got bled into the totals row.\n"
                "  3. Auto-detects per-pay-period exports (multiple rows per associate) and "
                "aggregates to one row per associate -- money/hours SUMmed, period dates "
                "MIN/MAX'd, identity columns kept as-is. Same-pay-date duplicates (real "
                "distinct paychecks in ADP) are also folded in by the SUM aggregation, which "
                "is correct for ADP.\n\n"
                "Optional swap: NET PAY <-> TAKE HOME values can be exchanged (default ON) "
                "because the Carvan-style API maps them reversed. Column headers are NEVER "
                "renamed -- only the data is swapped.\n\n"
                "Output: a CSV file with the input's exact column headers and column order, "
                "saved to the Audit Files inbox. Input accepts .xlsx / .xls / .csv. ADP money "
                "cells are stored as =ROUND(x, 2.0) Excel formulas; this tool evaluates them "
                "with openpyxl + a small formula parser, so values come through correctly."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "file_base64": {"type": "string", "description": "Fallback: base64 encoded ADP Prior Payroll file"},
                    "filename": {"type": "string", "description": "Optional filename hint, used for extension dispatch when file_base64 is supplied."},
                    "swap_net_take": {
                        "type": "boolean",
                        "description": "Swap NET PAY <-> TAKE HOME values (the API expects them reversed). Default true.",
                    },
                    "aggregation_strategy": {
                        "type": "string",
                        "enum": ["ask", "full_quarter", "preserve_pay_periods"],
                        "description": (
                            "How to handle multi-row-per-associate files.\n"
                            "  'ask' (DEFAULT) -- runs detection only, returns facts + a "
                            "recommendation (full_quarter vs preserve_pay_periods), and writes "
                            "NO file. Show the recommendation to the user, get their explicit "
                            "choice, then re-call this tool with the chosen value below. "
                            "Always start here unless the user has already told you which "
                            "strategy they want.\n"
                            "  'full_quarter' -- collapses all rows for an associate into one "
                            "(matches the Streamlit 'Full Quarter' radio). Right for full-"
                            "quarter per-pay-period exports the implementor mistakenly left "
                            "un-aggregated.\n"
                            "  'preserve_pay_periods' -- keeps distinct pay periods intact, "
                            "merging only same-day duplicate row pairs (matches the Streamlit "
                            "'Preserve Pay Periods' radio). Right for partial-quarter exports "
                            "where the API expects per-period rows."
                        ),
                    },
                },
            },
        ),
        types.Tool(
            name="adp_prior_payroll_setup_helper",
            description=(
                "[VENDOR: ADP only] [INPUT: SANITIZED ADP Prior Payroll file (.xlsx/.csv) -- "
                "run adp_prior_payroll_sanity first if the file has interleaved totals or "
                "per-pay-period rows] [SECONDARY INPUT: State Tax Code master CSV]\n"
                "[DO NOT USE FOR: UZIO files / Paycom files - the runtime guard will refuse.]\n\n"
                "Discovers what to configure in Uzio for an ADP Prior Payroll migration. "
                "Given a (sanitized) ADP Prior Payroll file plus the State Tax Code master "
                "CSV, produces an Excel workbook with:\n"
                "  - Earnings catalog (REGULAR / OVERTIME + every ADDITIONAL EARNINGS code)\n"
                "  - Contributions catalog (401k / 403b / 457 / Roth / HSA / FSA codes)\n"
                "  - Deductions catalog with pre-tax vs post-tax verdict per code\n"
                "  - Taxes discovered (every '* - EMPLOYEE/EMPLOYER TAX' column)\n"
                "  - Tax_Mapping sheet (and a CSV alongside) in the "
                "'Payroll_Mappings_Tax_Mapping_CORRECTED' format, with one row per "
                "(tax_type, state) - federal taxes get 1 row, state-scoped taxes (SIT, SDI, "
                "SUTA, FLI) get 1 row per distinct WORKED IN STATE present in the file\n"
                "  - Bonus_Classification (FLSA test: actual OT pay vs 1.5 x regular rate; "
                "any single row showing inflation = bonus is non-discretionary)\n\n"
                "Pre/post-tax algorithm: for each row, gap_FIT = TOTAL EARNINGS minus "
                "FEDERAL INCOME - EMPLOYEE TAXABLE. Find any subset of non-zero deductions "
                "summing to gap_FIT; every member is pre-tax for FIT. One positive proof in "
                "the whole file = pre-tax for everyone (the rule never varies per employee). "
                "Same logic for FICA / MEDI / SIT to derive flavor: section_125 (medical/"
                "dental/vision pre-FIT/FICA/MEDI/SIT) vs 401k_traditional (pre-FIT/SIT only)."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "file_base64": {"type": "string", "description": "Fallback: base64 encoded ADP Prior Payroll file (sanitized)."},
                    "filename": {"type": "string", "description": "Optional filename hint, used for extension dispatch when file_base64 is supplied."},
                    "state_tax_master_path": {
                        "type": "string",
                        "description": (
                            "Local path to the State Tax Code master CSV "
                            "(default: C:/Users/shobhit.sharma/Downloads/State Tax Code.csv). "
                            "Required to populate Uzio Tax Code / Unique Tax ID columns."
                        ),
                    },
                    "state_tax_master_base64": {
                        "type": "string",
                        "description": "Fallback: base64 encoded State Tax Code master CSV.",
                    },
                },
            },
        ),
        types.Tool(
            name="adp_census_generator",
            description=(
                "[VENDOR: ADP only] [INPUT: ADP Census export (.xlsx/.csv)]\n"
                "[OUTPUT: Uzio Census Template (.xlsm) -- the BLANK template is read from the "
                "templates folder and filled with the ADP data]\n"
                "[DO NOT USE FOR: UZIO files (this generates UZIO from ADP, not the reverse) "
                "or Paycom files (use 'paycom_census_generator'). Runtime guard enforces.]\n\n"
                "Generates a Uzio Census Template (.xlsm) from an ADP Census export. "
                "Reads ADP columns (e.g. 'Associate ID', 'Legal First Name', "
                "'FLSA Description'), maps them to the Uzio template's expected headers, "
                "applies the same auto-correction toggles as the Streamlit "
                "'ADP - Full Census Generation' tool, and writes the output as a .xlsm to "
                "the Audit Files inbox.\n\n"
                "WORKFLOW: copy the ADP file with copy_to_audit_inbox first, then pass the "
                "resulting path as 'file_path'. The output preserves the Uzio template's "
                "VBA macros, instructions, and all non-data sheets."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "file_base64": {"type": "string", "description": "Fallback: base64-encoded ADP Census export (use only if path unavailable)."},
                    "filename": {"type": "string", "description": "Optional filename hint, used when file_base64 is provided so the loader can pick the right reader (.xlsx vs .csv)."},
                    "fix_flsa": {"type": "boolean", "description": "Enforce FLSA/Pay Type alignment (Hourly→Non-Exempt, Salaried→Exempt; blank→Non-Exempt)."},
                    "fix_emails": {"type": "boolean", "description": "Use Personal Email when Work Email is blank."},
                    "fix_status": {"type": "boolean", "description": "Map Position Status to ACTIVE / TERMINATED / EXCLUDE; 'not hired' rows are dropped."},
                    "fix_inactive": {"type": "boolean", "description": "Map 'Inactive' to TERMINATED only when a Termination Date exists, else ACTIVE."},
                    "fix_type": {"type": "boolean", "description": "Map Worker Category to Full Time / Part Time / Seasonal / Other."},
                    "fix_zip": {"type": "boolean", "description": "Pad 4-digit zips and trim to 5 digits."},
                    "fix_license": {"type": "boolean", "description": "Clear license expiration when license number is missing or 00/00/0000."},
                    "fix_dol_status": {"type": "boolean", "description": "Default blank Employment Type to 'Full Time'."},
                },
            },
        ),
        types.Tool(
            name="paycom_census_generator",
            description=(
                "[VENDOR: Paycom only] [INPUT: Paycom Census export (.xlsx/.csv) with columns "
                "Employee_Code, Legal_Firstname, Exempt_Status, etc.]\n"
                "[OUTPUT: Uzio Census Template (.xlsm)]\n"
                "[DO NOT USE FOR: ADP files (use 'adp_census_generator') or UZIO files - "
                "the runtime guard will refuse non-Paycom inputs.]\n\n"
                "Generates a Uzio Census Template (.xlsm) from a Paycom Census export. "
                "Reads Paycom columns (e.g. 'Employee_Code', 'Legal_Firstname', "
                "'Exempt_Status'), maps them to the Uzio template's expected headers, "
                "applies the same auto-correction toggles as the Streamlit "
                "'Paycom - Full Census Generation' tool, and writes the output as a "
                ".xlsm to the Audit Files inbox.\n\n"
                "WORKFLOW: copy the Paycom file with copy_to_audit_inbox first, then "
                "pass the resulting path as 'file_path'. The output preserves the Uzio "
                "template's VBA macros, instructions, and all non-data sheets."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "file_base64": {"type": "string", "description": "Fallback: base64-encoded Paycom Census export (use only if path unavailable)."},
                    "filename": {"type": "string", "description": "Optional filename hint, used when file_base64 is provided."},
                    "fix_flsa": {"type": "boolean", "description": "Enforce FLSA/Pay Type alignment."},
                    "fix_emails": {"type": "boolean", "description": "Use Personal_Email when Work_Email is blank."},
                    "fix_status": {"type": "boolean", "description": "Map Employee_Status to ACTIVE / TERMINATED / EXCLUDE."},
                    "fix_inactive": {"type": "boolean", "description": "Map 'Inactive' to TERMINATED only when a Termination_Date exists, else ACTIVE."},
                    "fix_type": {"type": "boolean", "description": "Map DOL_Status to Full Time / Part Time / Seasonal / Other."},
                    "fix_position": {"type": "boolean", "description": "Auto-fill blank Job Title (Position) from Department_Desc."},
                    "fix_dol_status": {"type": "boolean", "description": "Default blank DOL_Status (Employment Type) to 'Full Time'."},
                    "fix_zip": {"type": "boolean", "description": "Pad 4-digit zips and trim to 5 digits."},
                    "fix_license": {"type": "boolean", "description": "Clear license expiration when DriversLicense is missing or 00/00/0000."},
                },
            },
        ),
        types.Tool(
            name="adp_emergency_audit",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [uzio_file_path: UZIO Emergency Contacts] "
                "[adp_file_path: ADP Emergency Contact Report]\n"
                "Audits emergency contact information (name, relationship, phone) between Uzio "
                "and ADP per employee."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "adp_raw_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="adp_license_audit",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [uzio_file_path: UZIO License Report] "
                "[adp_file_path: ADP License / Certification Report]\n"
                "Audits professional licenses (number, type, expiration) between Uzio and ADP."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "adp_raw_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="adp_timeoff_audit",
            description=(
                "[VENDOR: ADP + UZIO (two-file audit)] [uzio_file_path: UZIO Time Off Balances] "
                "[adp_file_path: ADP Accrual / PTO Balance Report]\n"
                "Audits time-off balances (PTO, sick, vacation accruals) between Uzio and ADP."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "adp_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "adp_raw_base64": {"type": "string"},
                },
            },
        ),

        # --- PAYCOM TOOLS ---
        types.Tool(
            name="paycom_selective_census_sync",
            description=(
                "Updates ONLY the requested columns in a pre-filled Uzio Census Template "
                "(.xlsm) using a fresh Paycom census export. Same shape as "
                "adp_selective_census_sync -- selected_uzio_cols are the Uzio raw mapping "
                "keys to overwrite, employees not present in the Paycom source are left "
                "untouched, and the .xlsm's VBA / instruction sheets / unselected columns "
                "all pass through unchanged.\n\n"
                "Job Title and Work Location: pass an explicit dict {source_value: "
                "uzio_value}, pass {} to seed automatically from the existing template, or "
                "omit/null to skip syncing those columns.\n\n"
                "Set discover_only=true to skip writing and just return the seed mappings + "
                "unique source values for caller review."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "paycom_file_path": {"type": "string", "description": PATH_DESC},
                    "paycom_file_base64": {"type": "string", "description": "Fallback: base64 Paycom census file"},
                    "filename": {"type": "string", "description": "Optional filename hint when using base64."},
                    "uzio_template_path": {"type": "string", "description": PATH_DESC + " (the pre-filled .xlsm)"},
                    "uzio_template_base64": {"type": "string", "description": "Fallback: base64 Uzio template"},
                    "selected_uzio_cols": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Keys from UZIO_RAW_MAPPING -- the columns to overwrite.",
                    },
                    "job_title_mapping": {
                        "type": "object",
                        "description": "Optional {source_job_title: uzio_job_title} dict. Pass {} to seed from template.",
                        "additionalProperties": {"type": "string"},
                    },
                    "work_location_mapping": {
                        "type": "object",
                        "description": "Optional {source_location: uzio_location} dict. Pass {} to seed from template.",
                        "additionalProperties": {"type": "string"},
                    },
                    "fix_options": {
                        "type": "object",
                        "description": "Optional auto-fix toggles.",
                        "additionalProperties": {"type": "boolean"},
                    },
                    "discover_only": {
                        "type": "boolean",
                        "description": "If true, return only the seed mappings without writing the template.",
                    },
                },
                "required": ["selected_uzio_cols"],
            },
        ),
        types.Tool(
            name="paycom_prior_payroll_generator",
            description=(
                "Generates a filled Uzio Prior Payroll Template (.xlsx) from 1-10 Paycom "
                "Prior Payroll files (long format with Type Code / Type Description / Code "
                "Description / Amount). Each (type_code, type_description) pair is auto-"
                "mapped to a Uzio target column via a fuzzy-string heuristic. "
                "'Net Pay Distribution' rows are auto-summed into the Uzio 'Net Pay' column; "
                "'Employee Benefits' rows are skipped. Pay-period dates are pulled from the "
                "filename pattern 'Pay Period MMDDYYYY MMDDYYYY Pay Date MMDDYYYY'.\n\n"
                "Records are aggregated per (employee, pay-period); a validation pass flags "
                "any employee-period where Gross - Taxes - Deductions != Net Pay. The auto-"
                "mapping can be overridden per-pair with override_mapping (key format "
                "'type_code|type_description', value is the Uzio column index, or a "
                "negative integer to skip).\n\n"
                "WORKFLOW: copy both the blank Uzio template and the Paycom file(s) to the "
                "Audit Files inbox first, then pass the resulting paths."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_template_path": {"type": "string", "description": PATH_DESC},
                    "uzio_template_base64": {"type": "string", "description": "Fallback: base64-encoded Uzio Prior Payroll Template (headers only)."},
                    "paycom_file_paths": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Local paths to Paycom Prior Payroll .xlsx files (max 10).",
                    },
                    "paycom_files_base64": {
                        "type": "array", "items": {"type": "string"},
                        "description": "Fallback: base64-encoded Paycom files (max 10).",
                    },
                    "override_mapping": {
                        "type": "object",
                        "description": (
                            "Optional {'type_code|type_description': uzio_col_idx} override. "
                            "Use a negative integer to force-skip a pair. Auto-guessed pairs "
                            "are kept for any (tc, td) not present in this object."
                        ),
                        "additionalProperties": {"type": "integer"},
                    },
                    "client_name": {"type": "string", "description": "Optional client name; used in the output filename."},
                },
            },
        ),
        types.Tool(
            name="paycom_prior_payroll_setup_helper",
            description=(
                "[VENDOR: Paycom only (TWO Paycom files required)] "
                "[prior_payroll_path: Paycom Prior Payroll Register, long format with columns "
                "EE Code, Type Code, Type Description, Amount, Code Description] "
                "[scheduled_deductions_path: Paycom Scheduled Deductions Report with columns "
                "Deduction Code, Deduction Desc, Tax Treatment]\n"
                "[DO NOT USE FOR: ADP files (use 'adp_prior_payroll_setup_helper'), "
                "UZIO files, or single-Paycom-file calls - both files are required for the "
                "complete analysis.]\n\n"
                "Replaces the deprecated 'paycom_deduction_analyzer' tool. Discovers what to "
                "configure in Uzio for a fresh Paycom prior payroll migration. Produces a 3-tab "
                "Excel workbook:\n"
                "  Tab 1 - What to Set Up (Earnings | Contributions | Deductions, codes only).\n"
                "  Tab 2 - Pre-Tax vs Post-Tax (read straight from the Tax Treatment column "
                "of the Scheduled Deductions report - 'B' = Section 125 pre-tax, 'H' = 401k "
                "traditional pre-tax, 'A' = post-tax).\n"
                "  Tab 3 - Bonus Verdict (FLSA discretionary vs non-discretionary). "
                "Strategy A+C: when both plain OT and Paycom's WOT (Weighted Overtime) lines "
                "exist for the same employee+period, compare them. Any positive WOT-vs-OT gap "
                "means Paycom rolled a bonus into the regular rate => non-discretionary. When "
                "the differential test cannot run (only WOT, only OT, or no bonus codes), the "
                "verdict is 'indeterminate' with a note to supply a Payroll Register Detail "
                "with hours."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "prior_payroll_path": {"type": "string", "description": PATH_DESC + " (Paycom Prior Payroll Register, long format)"},
                    "prior_payroll_base64": {"type": "string", "description": "Fallback: base64 Paycom Prior Payroll Register"},
                    "scheduled_deductions_path": {"type": "string", "description": PATH_DESC + " (Paycom Scheduled Deductions Report)"},
                    "scheduled_deductions_base64": {"type": "string", "description": "Fallback: base64 Paycom Scheduled Deductions Report"},
                },
            },
        ),
        types.Tool(
            name="paycom_total_comparison",
            description=(
                "[VENDOR: Paycom + UZIO (two-file audit)] [paycom_file_paths: 1+ Paycom Prior "
                "Payroll files] [uzio_file_path: UZIO Master / Custom Report]\n"
                "[DO NOT confuse the slots - Paycom file goes in paycom_file_paths, UZIO file "
                "goes in uzio_file_path.]\n\n"
                "Performs a complete payroll total comparison between Paycom and Uzio reports, "
                "producing up to 7 sheets: Full Comparison, Mismatches Only, Employee Mismatches, "
                "All Employee Details, Duplicate Pay Periods, Pay Stub Counts, and Tax Rate Verification.\n\n"
                "[MAPPINGS ARE REQUIRED FOR FULL OUTPUT] Without mappings, only Pay Stub Counts "
                "and Tax Rate Verification will populate - the four comparison sheets will be "
                "silently empty and not written to the Excel file. The Streamlit version asks "
                "the user to upload 4 mapping files (Earnings, Deductions, Contributions, Taxes) "
                "and you MUST do the equivalent here.\n\n"
                "[BEFORE CALLING THIS TOOL] If the user has not provided mapping files or "
                "mappings_json, STOP and ASK them: 'Do you have the 4 mapping files (Earnings, "
                "Deductions, Contributions, Taxes mapping files - the same ones the Streamlit "
                "tool asks for)? Please share their paths.' Do NOT proceed with an empty mappings "
                "list - the resulting report will be missing the comparison sheets and the user "
                "will think the tool is broken.\n\n"
                "HINT: Always call 'list_audit_files' first to identify the correct file paths."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "paycom_file_paths": {"type": "array", "items": {"type": "string"}, "description": PATH_DESC},
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "mapping_file_paths": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": (
                            "REQUIRED for full report. Local paths to the 4 mapping files: "
                            "Earnings, Deductions, Contributions, and Taxes. If the user has not "
                            "provided these, ASK before calling - do NOT call this tool with an empty list."
                        ),
                    },
                    "mappings_json": {
                        "type": "string",
                        "description": (
                            "Alternative to mapping_file_paths. Flat JSON array of objects with "
                            "keys Category ('Earnings' | 'Deductions' | 'Contributions' | 'Taxes'), "
                            "ADP_Name (or Paycom code), UZIO_Name. Use only if the user provides mappings inline."
                        ),
                    },
                    "paycom_files_base64": {"type": "array", "items": {"type": "string"}},
                    "uzio_file_base64": {"type": "string"},
                },
                "required": [],
            },
        ),
        types.Tool(
            name="paycom_deduction_audit",
            description=(
                "[VENDOR: Paycom + UZIO (two-file audit)] [uzio_file_path: UZIO Deduction Report] "
                "[paycom_file_path: Paycom Deduction Register]\n"
                "[REQUIRED: mapping_json - JSON object mapping Paycom Deduction Description to "
                "UZIO Deduction Name]\n"
                "Compares deduction amounts between Uzio and Paycom per employee, per code."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "paycom_file_path": {"type": "string", "description": PATH_DESC},
                    "mapping_json": {"type": "string", "description": "JSON mapping of Paycom Deduction Description -> Uzio Deduction Name"},
                    "uzio_raw_base64": {"type": "string"},
                    "paycom_raw_base64": {"type": "string"},
                },
                "required": ["mapping_json"],
            },
        ),
        types.Tool(
            name="paycom_census_audit",
            description=(
                "[VENDOR: Paycom + UZIO (two-file audit)] [uzio_file_path: UZIO Census Custom Report] "
                "[paycom_file_path: Paycom Census export (.xlsx/.csv)]\n"
                "Audits employee census data between Uzio and Paycom to find mismatches in names, "
                "emails, addresses, hire/term dates, status, FLSA, etc."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "paycom_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string", "description": "Fallback: base64 Uzio file"},
                    "paycom_raw_base64": {"type": "string", "description": "Fallback: base64 Paycom file"},
                },
            },
        ),
        types.Tool(
            name="paycom_sql_master",
            description=(
                "[VENDOR: Paycom only] [INPUT: Paycom UPS SQL Master file (.sql/.csv/.xlsx)]\n"
                "[DO NOT USE FOR: ADP files, UZIO files, or generic SQL exports - this is a "
                "Paycom-specific consolidation report.]\n"
                "Processes a Paycom UPS SQL Master file into a structured audit report."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "sql_file_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="paycom_payment_audit",
            description=(
                "[VENDOR: Paycom + UZIO (two-file audit)] [uzio_file_path: UZIO Payment Report] "
                "[paycom_file_path: Paycom Payment Register]\n"
                "Audits payment methods (direct deposit / check) and amounts between Uzio and Paycom."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "paycom_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "paycom_raw_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="paycom_emergency_audit",
            description=(
                "[VENDOR: Paycom + UZIO (two-file audit)] [uzio_file_path: UZIO Emergency Contacts] "
                "[paycom_file_path: Paycom Emergency Contact Report]\n"
                "Audits emergency contact information (name, relationship, phone) between Uzio and Paycom."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "paycom_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "paycom_raw_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="paycom_timeoff_audit",
            description=(
                "[VENDOR: Paycom + UZIO (two-file audit)] [uzio_file_path: UZIO Time Off Balances] "
                "[paycom_file_path: Paycom Accrual / PTO Balance Report]\n"
                "Audits time-off balances (PTO, sick, vacation accruals) between Uzio and Paycom."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "paycom_file_path": {"type": "string", "description": PATH_DESC},
                    "uzio_raw_base64": {"type": "string"},
                    "paycom_raw_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="paycom_withholding_audit",
            description=(
                "[VENDOR: Paycom + UZIO (two-file audit)] [uzio_file_path: UZIO Withholding Report] "
                "[paycom_file_path: Paycom Withholding / W-4 Report]\n"
                "Audits federal + state tax withholding settings (filing status, allowances, "
                "extra withholding, exemptions) between Uzio and Paycom per employee."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC},
                    "paycom_file_path": {"type": "string", "description": PATH_DESC},
                    "mapping_file_path": {"type": "string"},
                    "uzio_raw_base64": {"type": "string"},
                    "paycom_raw_base64": {"type": "string"},
                    "mapping_file_base64": {"type": "string"},
                },
            },
        ),
        types.Tool(
            name="paycom_census_sanity",
            description=(
                "[VENDOR: Paycom only] [INPUT: Paycom Census export (.xlsx/.csv) with columns "
                "Employee_Code, SS_Number, Department_Desc, DOL_Status, etc.]\n"
                "[DO NOT USE FOR: ADP Census (use 'adp_census_sanity') or UZIO files - the "
                "runtime guard will refuse non-Paycom inputs.]\n\n"
                "Applies opt-in auto-corrections to a Paycom Census export. "
                "MANDATORY: For stability, always use copy_to_audit_inbox first and then use 'file_path'. "
                "Do NOT use 'file_base64' for files > 1MB."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "file_base64": {"type": "string", "description": "Fallback: base64 encoded Paycom Census export"},
                    "filename": {"type": "string"},
                    "fix_flsa": {"type": "boolean", "description": "Enforce FLSA/Pay Type alignment — fill blank FLSA from Hourly/Salaried."},
                    "fix_emails": {"type": "boolean", "description": "Use Personal Email as fallback when Work Email is blank."},
                    "fix_driver_smart": {"type": "boolean", "description": "Smart Driver Correction: if Dept/Job indicates Driver, fill Job/FLSA/Pay Type."},
                    "fix_license": {"type": "boolean", "description": "Strict license validation — clear license dates if number is missing."},
                    "fix_status": {"type": "boolean", "description": "Auto-map Employment Status (e.g. Inactive -> Terminated)."},
                    "fix_type": {"type": "boolean", "description": "Auto-map Worker Category (e.g. Intern -> Part Time)."},
                    "fix_position": {"type": "boolean", "description": "Auto-Fill blank Position with Department Description."},
                    "fix_dol_status": {"type": "boolean", "description": "Auto-fill blank DOL_Status to 'Full-Time' for active employees."},
                    "fix_zip": {"type": "boolean", "description": "Auto-fix Zip Code (pad 4-digits and trim to 5-digits)."},
                    "sort_by_manager": {"type": "boolean", "description": "Cluster managers and their reportees at the top of the output."},
                },
            },
        ),
        types.Tool(
            name="paycom_consolidated_audit",
            description=(
                "Runs Census + Payment + Emergency contact audits in one pass against the "
                "Uzio Master Custom Report (CSV with category labels in row 1, headers in "
                "row 2) and a Paycom Census export (.xlsx or .csv). Produces a single "
                "consolidated report with these sheets: Summary (per-metric counts), "
                "Duplicate_SSN_Check, Census_Audit, Payment_Audit, Emergency_Audit, plus "
                "anomaly extracts -- Salaried_Drivers, FLSA_Issues, Active_Missing, "
                "Terminated_Missing, Data_Quality, High_Rate_Anomalies.\n\n"
                "Identity matching is done by Employee ID with SSN as a fallback; the "
                "audit honors the same canonical comparisons used by the per-vendor "
                "audits (pay-type aware, termination-reason flexible matching, "
                "phone/zip/SSN/date/money normalizers).\n\n"
                "WORKFLOW: copy both files to the Audit Files inbox first, then pass the "
                "resulting paths."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "uzio_file_path": {"type": "string", "description": PATH_DESC + " (Uzio Master Custom Report CSV)"},
                    "paycom_file_path": {"type": "string", "description": PATH_DESC + " (Paycom Census Export xlsx/csv)"},
                    "uzio_raw_base64": {"type": "string", "description": "Fallback: base64-encoded Uzio Master CSV"},
                    "paycom_raw_base64": {"type": "string", "description": "Fallback: base64-encoded Paycom Census Export"},
                    "client_name": {"type": "string", "description": "Optional client name; used in the output filename."},
                },
            },
        ),
        types.Tool(
            name="selective_employee_extractor",
            description=(
                "[VENDOR-AGNOSTIC] [INPUT: any payroll/census file (ADP, Paycom, UZIO) with "
                "an Employee ID column] [OUTPUT: a slimmed-down file containing only the "
                "rows for the requested IDs]\n"
                "Use this to isolate problematic employees flagged by an audit, so downstream "
                "tools work on a manageable subset. Preserves the input file's schema."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "employee_ids": {"type": "array", "items": {"type": "string"}, "description": "List of Employee IDs to extract"},
                    "file_base64": {"type": "string", "description": "Fallback: base64 encoded file"},
                },
                "required": ["employee_ids"],
            },
        ),
        types.Tool(
            name="read_audit_report",
            description=(
                "[VENDOR-AGNOSTIC] [INPUT: any Excel or CSV file]\n"
                "Reads the full contents of an audit report (Excel or CSV) into the response. "
                "Use this AFTER an audit tool has produced output, to inspect specific rows. "
                "WARNING: Loads the whole file into context; for files > 1MB use get_file_schema "
                "+ query_data_sql instead."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "sheet_name": {"type": "string", "description": "Optional: Name of the sheet to read (defaults to first sheet)"},
                },
                "required": ["file_path"],
            },
        ),

        # --- DATABASE & QUERY TOOLS ---
        types.Tool(
            name="get_file_schema",
            description=(
                "[VENDOR-AGNOSTIC] [INPUT: any CSV or Excel file]\n"
                "Inspects a file and returns column names, dtypes, and a 5-row sample. "
                "Caches the file in memory for 5x faster subsequent query_data_sql calls. "
                "Returns 'date_hint_columns' so you know which columns need CAST(... AS DATE) "
                "in SQL. ALWAYS call this before query_data_sql so you know the schema."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "sheet_name": {"type": "string", "description": "Optional: Excel sheet name (defaults to auto-detected/first)"},
                },
                "required": ["file_path"],
            },
        ),
        types.Tool(
            name="query_data_sql",
            description=(
                "[VENDOR-AGNOSTIC] [INPUT: any CSV or Excel file]\n"
                "Executes a SQL query against the file using DuckDB. The file is loaded into "
                "a single in-memory table named 'data'. Use this for files too big to fit in "
                "context (the 1MB+ problem). The connection is ephemeral; DDL statements are "
                "discarded between calls.\n"
                "PRECONDITION: call get_file_schema first to learn column names and dtypes. "
                "Date columns often arrive as type 'object' (string), so wrap with "
                "CAST(col AS DATE) when filtering by date."
            ),
            inputSchema={
                "type": "object",
                "properties": {
                    "file_path": {"type": "string", "description": PATH_DESC},
                    "sql_query": {"type": "string", "description": "The SQL query to run. Use 'data' as the table name."},
                    "sheet_name": {"type": "string", "description": "Optional: Excel sheet name."},
                },
                "required": ["file_path", "sql_query"],
            },
        ),
    ]

# ── Tool Handlers ─────────────────────────────────────────────────────────────

@server.call_tool()
async def handle_call_tool(name: str, arguments: dict | None):
    arguments = arguments or {}
    try:
        if name == "list_audit_files":
            files = []
            scan_dir = arguments.get("directory_path", AUDIT_INBOX).strip().strip('"')
            
            if not os.path.isdir(scan_dir):
                return [types.TextContent(type="text", text=f"Error: The directory '{scan_dir}' does not exist or is not a folder.")]

            try:
                from datetime import datetime
                count = 0
                for root, dirs, filenames in os.walk(scan_dir):
                    for fname in sorted(filenames):
                        if count > 500: break # Safety limit
                        fpath = os.path.join(root, fname)
                        try:
                            stat = os.stat(fpath)
                            files.append({
                                "name": fname,
                                "path": fpath,
                                "size_kb": round(stat.st_size / 1024, 1),
                                "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                            })
                            count += 1
                        except: continue
                    if count > 500: break
            except Exception as e:
                return [types.TextContent(type="text", text=f"Error scanning directory: {str(e)}")]

            result = {
                "scanned_directory": scan_dir,
                "file_count": len(files),
                "files": files,
                "note": "Limited to first 500 files for performance." if len(files) >= 500 else "All files listed.",
                "instruction": "Use the 'path' field from these files as arguments in other tools.",
            }
            return [types.TextContent(type="text", text=json.dumps(result, indent=2, default=_json_default))]

        elif name == "copy_to_audit_inbox":
            source = arguments.get("source_path")
            result = copy_file_to_inbox(source)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2, default=_json_default))]

        elif name == "apply_data_corrections":
            path = arguments.get("file_path")
            corrs = arguments.get("corrections", [])
            result = apply_data_corrections(path, corrs)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2, default=_json_default))]

        elif name == "adp_total_comparison":
            adp_data = load_files_list(arguments, "adp_file_paths", "adp_files_base64")
            uzio_data = (load_file(arguments, "uzio_file_path", "uzio_file_base64"), "uzio.xlsx")
            
            mapping_paths = arguments.get("mapping_file_paths", [])
            if mapping_paths:
                mappings = load_mappings_from_paths(mapping_paths)
            else:
                mappings_raw = arguments.get("mappings_json", "[]")
                try:
                    mappings = json.loads(mappings_raw)
                    if isinstance(mappings, dict):
                        # Flatten if passed as a dict of categories
                        flattened = []
                        for cat, items in mappings.items():
                            if isinstance(items, list):
                                for item in items:
                                    if isinstance(item, dict):
                                        item["Category"] = item.get("Category", cat)
                                        flattened.append(item)
                        mappings = flattened
                except:
                    return [types.TextContent(type="text", text="Error: mappings_json is not valid JSON.")]

            results = run_adp_total_comparison(adp_data, uzio_data, mappings)
            summary = save_results_to_excel(results, "ADP_Total_Comparison")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "adp_census_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            adp = load_file(arguments, "adp_file_path", "adp_raw_base64")
            results = run_adp_census_audit(uzio, adp)
            summary = save_results_to_excel(results, "ADP_Census_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "adp_deduction_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            adp = load_file(arguments, "adp_file_path", "adp_raw_base64")
            mapping = json.loads(arguments.get("mapping_json", "{}"))
            results = run_adp_deduction_audit(uzio, adp, mapping)
            summary = save_results_to_excel(results, "ADP_Deduction_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "adp_payment_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            adp = load_file(arguments, "adp_file_path", "adp_raw_base64")
            results = run_adp_payment_audit(uzio, adp)
            summary = save_results_to_excel(results, "ADP_Payment_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "adp_withholding_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            adp = load_file(arguments, "adp_file_path", "adp_raw_base64")
            results = run_adp_withholding_audit(uzio, adp)
            summary = save_results_to_excel(results, "ADP_Withholding_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "adp_census_sanity":
            content = load_file(arguments, "file_path", "file_base64")
            filename = arguments.get("filename") or (
                os.path.basename(arguments["file_path"].strip().strip('"'))
                if arguments.get("file_path") else "upload.xlsx"
            )
            require_vendor(content, filename, "adp", "adp_census_sanity")
            toggle_keys = [
                "fix_flsa", "fix_emails", "fix_job_title", "fix_driver_smart", "fix_license",
                "fix_status", "fix_type", "fix_dol_status", "fix_leave_to_active",
                "fix_blank_jt_to_driver", "fix_std_hours", "rename_std_hours",
                "fix_zip", "rename_zip_col", "replace_gender_col",
            ]
            fix_options = {k: bool(arguments.get(k, False)) for k in toggle_keys}
            fix_options["fix_inactive"] = fix_options["fix_status"]
            sort_by_manager = bool(arguments.get("sort_by_manager", False))
            xlsx_bytes, summary = generate_corrected_census_xlsx(
                content, ADP_FIELD_MAP, fix_options=fix_options,
                filename=filename, sort_by_manager=sort_by_manager,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            # Save output to Audit Files folder for consistency
            out_path = os.path.join(AUDIT_INBOX, f"ADP_Cleaned_{stamp}.xlsx")
            with open(out_path, "wb") as f:
                f.write(xlsx_bytes)
            payload = {
                "output_file": out_path,
                "summary": summary,
                "applied_toggles": {k: v for k, v in fix_options.items() if v},
            }
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "paycom_consolidated_audit":
            uzio_content = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            paycom_content = load_file(arguments, "paycom_file_path", "paycom_raw_base64")
            paycom_path = arguments.get("paycom_file_path")
            paycom_filename = (
                os.path.basename(paycom_path) if paycom_path else "paycom.xlsx"
            )
            client_name = (arguments.get("client_name") or "").strip()
            results = run_paycom_consolidated_audit(uzio_content, paycom_content, paycom_filename)
            prefix = f"Paycom_Consolidated_Audit_{client_name}".rstrip("_") if client_name else "Paycom_Consolidated_Audit"
            summary = save_results_to_excel(results, prefix)
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_selective_census_sync":
            paycom_content = load_file(arguments, "paycom_file_path", "paycom_file_base64")
            uzio_content = load_file(arguments, "uzio_template_path", "uzio_template_base64")
            paycom_path = arguments.get("paycom_file_path")
            filename = arguments.get("filename") or (
                os.path.basename(paycom_path) if paycom_path else "census.xlsx"
            )
            selected = arguments.get("selected_uzio_cols") or []
            job_map = arguments.get("job_title_mapping")
            loc_map = arguments.get("work_location_mapping")
            fix_options = arguments.get("fix_options") or {}

            if arguments.get("discover_only"):
                info = paycom_selective_discover(paycom_content, filename, uzio_content)
                return [types.TextContent(type="text", text=json.dumps(info, indent=2, default=_json_default))]

            xlsm_bytes, summary = run_paycom_selective_census_sync(
                paycom_content, filename, uzio_content,
                selected_uzio_cols=selected,
                job_title_mapping=job_map,
                work_location_mapping=loc_map,
                fix_options=fix_options,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            out_name = f"Uzio_Updated_Paycom_{stamp}.xlsm"
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            out_path = os.path.join(AUDIT_INBOX, out_name)
            with open(out_path, "wb") as f:
                f.write(xlsm_bytes)
            payload = {"output_file": out_path, "summary": summary}
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "paycom_prior_payroll_generator":
            uzio_bytes = load_file(arguments, "uzio_template_path", "uzio_template_base64")
            paycom_files = load_files_list(arguments, "paycom_file_paths", "paycom_files_base64")
            if not paycom_files:
                return [types.TextContent(type="text", text="Error: provide paycom_file_paths or paycom_files_base64.")]
            if len(paycom_files) > 10:
                return [types.TextContent(type="text", text="Error: maximum 10 Paycom files supported.")]
            override_mapping = arguments.get("override_mapping") or None
            client_name = (arguments.get("client_name") or "").strip()
            xlsx_bytes, summary = run_paycom_prior_payroll_generator(
                uzio_bytes, paycom_files, override_mapping=override_mapping,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            base = (client_name + "_" if client_name else "")
            out_name = f"Uzio_Prior_Payroll_Paycom_{base}{stamp}.xlsx".replace(" ", "_")
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            out_path = os.path.join(AUDIT_INBOX, out_name)
            with open(out_path, "wb") as f:
                f.write(xlsx_bytes)
            payload = {"output_file": out_path, "summary": summary}
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "adp_selective_census_sync":
            adp_content = load_file(arguments, "adp_file_path", "adp_file_base64")
            uzio_content = load_file(arguments, "uzio_template_path", "uzio_template_base64")
            adp_path = arguments.get("adp_file_path")
            filename = arguments.get("filename") or (
                os.path.basename(adp_path) if adp_path else "census.xlsx"
            )
            selected = arguments.get("selected_uzio_cols") or []
            job_map = arguments.get("job_title_mapping")
            loc_map = arguments.get("work_location_mapping")
            fix_options = arguments.get("fix_options") or {}

            if arguments.get("discover_only"):
                info = adp_selective_discover(adp_content, filename, uzio_content)
                return [types.TextContent(type="text", text=json.dumps(info, indent=2, default=_json_default))]

            xlsm_bytes, summary = run_adp_selective_census_sync(
                adp_content, filename, uzio_content,
                selected_uzio_cols=selected,
                job_title_mapping=job_map,
                work_location_mapping=loc_map,
                fix_options=fix_options,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            out_name = f"Uzio_Updated_ADP_{stamp}.xlsm"
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            out_path = os.path.join(AUDIT_INBOX, out_name)
            with open(out_path, "wb") as f:
                f.write(xlsm_bytes)
            payload = {"output_file": out_path, "summary": summary}
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "adp_prior_payroll_generator":
            uzio_bytes = load_file(arguments, "uzio_template_path", "uzio_template_base64")
            adp_files = load_files_list(arguments, "adp_file_paths", "adp_files_base64")
            if not adp_files:
                return [types.TextContent(type="text", text="Error: provide adp_file_paths or adp_files_base64.")]
            if len(adp_files) > 10:
                return [types.TextContent(type="text", text="Error: maximum 10 ADP files supported.")]
            override_mapping = arguments.get("override_mapping") or None
            client_name = (arguments.get("client_name") or "").strip()
            xlsx_bytes, summary = run_adp_prior_payroll_generator(
                uzio_bytes, adp_files, override_mapping=override_mapping,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            base = (client_name + "_" if client_name else "")
            out_name = f"Uzio_Prior_Payroll_{base}{stamp}.xlsx".replace(" ", "_")
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            out_path = os.path.join(AUDIT_INBOX, out_name)
            with open(out_path, "wb") as f:
                f.write(xlsx_bytes)
            payload = {"output_file": out_path, "summary": summary}
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "adp_prior_payroll_sanity":
            content = load_file(arguments, "file_path", "file_base64")
            file_path_arg = arguments.get("file_path")
            filename = arguments.get("filename") or (
                os.path.basename(file_path_arg) if file_path_arg else "upload.xlsx"
            )
            require_vendor(content, filename, "adp", "adp_prior_payroll_sanity")
            swap_net_take = bool(arguments.get("swap_net_take", True))
            agg_strategy = arguments.get("aggregation_strategy") or "ask"
            csv_bytes, summary = run_adp_prior_payroll_sanity(
                content,
                filename=filename,
                swap_net_take=swap_net_take,
                aggregation_strategy=agg_strategy,
            )
            if summary.get("mode") == "detection_only":
                # Detection-only mode: no file written, return facts + recommendation.
                return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            base = os.path.splitext(filename)[0] or "ADP_Prior_Payroll"
            out_name = f"{base}_Sanity_Cleaned_{stamp}.csv"
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            out_path = os.path.join(AUDIT_INBOX, out_name)
            with open(out_path, "wb") as f:
                f.write(csv_bytes)
            payload = {
                "output_file": out_path,
                "summary": summary,
                "swap_applied": summary.get("swap_applied", False),
            }
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "adp_prior_payroll_setup_helper":
            content = load_file(arguments, "file_path", "file_base64")
            file_path_arg = arguments.get("file_path")
            filename = arguments.get("filename") or (
                os.path.basename(file_path_arg) if file_path_arg else "adp_prior_payroll.xlsx"
            )
            require_vendor(content, filename, "adp", "adp_prior_payroll_setup_helper")
            master_path = arguments.get("state_tax_master_path") or r"C:\Users\shobhit.sharma\Downloads\State Tax Code.csv"
            master_b64 = arguments.get("state_tax_master_base64")
            master_content = b""
            if master_path and os.path.isfile(master_path.strip().strip('"')):
                with open(master_path.strip().strip('"'), "rb") as f:
                    master_content = f.read()
            elif master_b64:
                master_content = base64.b64decode(master_b64)
            results, csv_bytes = run_adp_prior_payroll_setup_helper(
                content, adp_filename=filename, state_tax_master_content=master_content,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            base = os.path.splitext(filename)[0] or "ADP_Prior_Payroll"
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            tax_csv_path = os.path.join(AUDIT_INBOX, f"{base}_Tax_Mapping_{stamp}.csv")
            with open(tax_csv_path, "wb") as f:
                f.write(csv_bytes)
            xlsx_path = os.path.join(AUDIT_INBOX, f"{base}_Setup_Helper_{stamp}.xlsx")
            with open(xlsx_path, "wb") as f:
                f.write(_setup_helper_xlsx(results))
            bonus = results["Bonus_Classification"][0]
            summary_info = {
                "output_file": xlsx_path,
                "tax_mapping_csv": tax_csv_path,
                "message": (
                    f"Setup helper produced a 3-tab xlsx in 'Audit Files'. "
                    f"Tab 1 = What to Set Up, Tab 2 = Bonus Verdict ({bonus['Verdict']}), "
                    f"Tab 3 = Pre-Tax vs Post-Tax per deduction."
                ),
            }
            return [types.TextContent(type="text", text=json.dumps(summary_info, indent=2, default=_json_default))]

        elif name == "adp_census_generator":
            content = load_file(arguments, "file_path", "file_base64")
            filename = arguments.get("filename") or os.path.basename(
                (arguments.get("file_path") or "adp_census.xlsx").strip().strip('"')
            )
            require_vendor(content, filename, "adp", "adp_census_generator")
            toggle_keys = [
                "fix_flsa", "fix_emails", "fix_status", "fix_inactive",
                "fix_type", "fix_zip", "fix_license", "fix_dol_status",
            ]
            fix_options = {k: bool(arguments.get(k, False)) for k in toggle_keys}
            xlsm_bytes, summary = run_adp_census_generation(content, filename, fix_options=fix_options)
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            out_path = os.path.join(AUDIT_INBOX, f"ADP_Uzio_Census_{stamp}.xlsm")
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            with open(out_path, "wb") as f:
                f.write(xlsm_bytes)
            payload = {
                "output_file": out_path,
                "summary": summary,
                "applied_toggles": {k: v for k, v in fix_options.items() if v},
                "message": f"Uzio Census Template written to {out_path}",
            }
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "paycom_census_generator":
            content = load_file(arguments, "file_path", "file_base64")
            filename = arguments.get("filename") or os.path.basename(
                (arguments.get("file_path") or "paycom_census.xlsx").strip().strip('"')
            )
            require_vendor(content, filename, "paycom", "paycom_census_generator")
            toggle_keys = [
                "fix_flsa", "fix_emails", "fix_status", "fix_inactive",
                "fix_type", "fix_position", "fix_dol_status", "fix_zip", "fix_license",
            ]
            fix_options = {k: bool(arguments.get(k, False)) for k in toggle_keys}
            xlsm_bytes, summary = run_paycom_census_generation(content, filename, fix_options=fix_options)
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            out_path = os.path.join(AUDIT_INBOX, f"Paycom_Uzio_Census_{stamp}.xlsm")
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            with open(out_path, "wb") as f:
                f.write(xlsm_bytes)
            payload = {
                "output_file": out_path,
                "summary": summary,
                "applied_toggles": {k: v for k, v in fix_options.items() if v},
                "message": f"Uzio Census Template written to {out_path}",
            }
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "adp_emergency_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            adp = load_file(arguments, "adp_file_path", "adp_raw_base64")
            results = run_adp_emergency_audit(uzio, adp)
            summary = save_results_to_excel(results, "ADP_Emergency_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "adp_license_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            adp = load_file(arguments, "adp_file_path", "adp_raw_base64")
            results = run_adp_license_audit(uzio, adp)
            summary = save_results_to_excel(results, "ADP_License_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "adp_timeoff_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            adp = load_file(arguments, "adp_file_path", "adp_raw_base64")
            results = run_adp_timeoff_audit(uzio, adp)
            # Timeoff usually returns a message if it's a template update
            if isinstance(results, dict) and "message" in results:
                return [types.TextContent(type="text", text=json.dumps(results, indent=2, default=_json_default))]
            summary = save_results_to_excel(results, "ADP_Timeoff_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_prior_payroll_setup_helper":
            prior = load_file(arguments, "prior_payroll_path", "prior_payroll_base64")
            sched = load_file(arguments, "scheduled_deductions_path", "scheduled_deductions_base64")
            prior_name = arguments.get("prior_payroll_path") or "paycom_prior_payroll.xlsx"
            sched_name = arguments.get("scheduled_deductions_path") or "paycom_scheduled.xlsx"
            prior_name = os.path.basename(prior_name.strip().strip('"'))
            sched_name = os.path.basename(sched_name.strip().strip('"'))
            require_vendor(prior, prior_name, "paycom", "paycom_prior_payroll_setup_helper (prior payroll)")
            require_vendor(sched, sched_name, "paycom", "paycom_prior_payroll_setup_helper (scheduled deductions)")
            results, xlsx_bytes = run_paycom_prior_payroll_setup_helper(
                prior, prior_name, sched, sched_name,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            base = os.path.splitext(prior_name)[0] or "Paycom_Prior_Payroll"
            if not os.path.exists(AUDIT_INBOX):
                os.makedirs(AUDIT_INBOX, exist_ok=True)
            out_path = os.path.join(AUDIT_INBOX, f"{base}_Setup_Helper_{stamp}.xlsx")
            with open(out_path, "wb") as f:
                f.write(xlsx_bytes)
            bonus = results["Bonus"]
            payload = {
                "output_file": out_path,
                "message": (
                    f"Paycom setup helper produced a 3-tab xlsx in 'Audit Files'. "
                    f"Tab 1 = What to Set Up, Tab 2 = Pre-Tax vs Post-Tax, "
                    f"Tab 3 = Bonus Verdict ({bonus['verdict']})."
                ),
                "answers": {
                    "earnings_to_set_up": [f"{r['Type Code']} - {r['Type Description']}"
                                            for r in results["Earnings_Codes"]],
                    "contributions_to_set_up": [f"{r['Deduction Code']} - {r['Deduction Desc']}"
                                                 for r in results["Contributions"]],
                    "deductions_to_set_up": [f"{r['Deduction Code']} - {r['Deduction Desc']}"
                                              for r in results["Deductions"]],
                    "pre_post_tax": [
                        {"code": r["Code"], "verdict": r["Verdict"], "flavor": r["Flavor"]}
                        for r in results["Pre_Post_Tax"]
                    ],
                    "bonus_verdict": bonus["verdict"],
                    "bonus_reason": bonus["reason"],
                },
            }
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "paycom_total_comparison":
            paycom_data = load_files_list(arguments, "paycom_file_paths", "paycom_files_base64")
            uzio_data = (load_file(arguments, "uzio_file_path", "uzio_file_base64"), "uzio.xlsx")
            
            mapping_paths = arguments.get("mapping_file_paths", [])
            if mapping_paths:
                mappings = load_mappings_from_paths(mapping_paths)
            else:
                mappings_raw = arguments.get("mappings_json", "[]")
                try:
                    mappings = json.loads(mappings_raw)
                    if isinstance(mappings, dict):
                        # Flatten if passed as a dict of categories
                        flattened = []
                        for cat, items in mappings.items():
                            if isinstance(items, list):
                                for item in items:
                                    if isinstance(item, dict):
                                        item["Category"] = item.get("Category", cat)
                                        flattened.append(item)
                        mappings = flattened
                except:
                    return [types.TextContent(type="text", text="Error: mappings_json is not valid JSON.")]

            results = run_paycom_total_comparison(paycom_data, uzio_data, mappings)
            summary = save_results_to_excel(results, "Paycom_Total_Comparison")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_deduction_audit":
            uzio_data = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            paycom_data = load_file(arguments, "paycom_file_path", "paycom_raw_base64")
            mapping = json.loads(arguments.get("mapping_json", "{}"))
            results = run_paycom_deduction_audit(uzio_data, paycom_data, mapping)
            summary = save_results_to_excel(results, "Paycom_Deduction_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_census_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            paycom = load_file(arguments, "paycom_file_path", "paycom_raw_base64")
            results = run_paycom_census_audit(uzio, paycom)
            summary = save_results_to_excel(results, "Paycom_Census_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_sql_master":
            content = load_file(arguments, "file_path", "sql_file_base64")
            results = run_paycom_sql_master(content)
            summary = save_results_to_excel(results, "Paycom_SQL_Master")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_payment_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            paycom = load_file(arguments, "paycom_file_path", "paycom_raw_base64")
            results = run_paycom_payment_audit(uzio, paycom)
            summary = save_results_to_excel(results, "Paycom_Payment_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_emergency_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            paycom = load_file(arguments, "paycom_file_path", "paycom_raw_base64")
            results = run_paycom_emergency_audit(uzio, paycom)
            summary = save_results_to_excel(results, "Paycom_Emergency_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_timeoff_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            paycom = load_file(arguments, "paycom_file_path", "paycom_raw_base64")
            results = run_paycom_timeoff_audit(uzio, paycom)
            summary = save_results_to_excel(results, "Paycom_Timeoff_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_withholding_audit":
            uzio = load_file(arguments, "uzio_file_path", "uzio_raw_base64")
            paycom = load_file(arguments, "paycom_file_path", "paycom_raw_base64")
            mapping = load_file(arguments, "mapping_file_path", "mapping_file_base64") or None
            results = run_paycom_withholding_audit(uzio, paycom, mapping)
            summary = save_results_to_excel(results, "Paycom_Withholding_Audit")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "paycom_census_sanity":
            content = load_file(arguments, "file_path", "file_base64")
            filename = arguments.get("filename") or (
                os.path.basename(arguments["file_path"].strip().strip('"'))
                if arguments.get("file_path") else "upload.xlsx"
            )
            require_vendor(content, filename, "paycom", "paycom_census_sanity")
            toggle_keys = [
                "fix_flsa", "fix_emails", "fix_driver_smart", "fix_license",
                "fix_status", "fix_type", "fix_position", "fix_dol_status", "fix_zip",
            ]
            fix_options = {k: bool(arguments.get(k, False)) for k in toggle_keys}
            fix_options["fix_inactive"] = fix_options["fix_status"]
            fix_options["fix_job_title"] = fix_options["fix_position"]
            sort_by_manager = bool(arguments.get("sort_by_manager", False))
            xlsx_bytes, summary = generate_corrected_census_xlsx(
                content, PAYCOM_FIELD_MAP, fix_options=fix_options,
                filename=filename, sort_by_manager=sort_by_manager,
            )
            from datetime import datetime
            stamp = datetime.now().strftime("%Y%m%d_%H%M")
            # Save output to Audit Files folder for consistency
            out_path = os.path.join(AUDIT_INBOX, f"Paycom_Cleaned_{stamp}.xlsx")
            with open(out_path, "wb") as f:
                f.write(xlsx_bytes)
            payload = {
                "output_file": out_path,
                "summary": summary,
                "applied_toggles": {k: v for k, v in fix_options.items() if v},
            }
            return [types.TextContent(type="text", text=json.dumps(payload, indent=2, default=_json_default))]

        elif name == "selective_employee_extractor":
            content = load_file(arguments, "file_path", "file_base64")
            import pandas as pd, io
            from utils.audit_utils import norm_id
            
            # Load file (try Excel then CSV)
            try: df = pd.read_excel(io.BytesIO(content), dtype=str)
            except: df = pd.read_csv(io.BytesIO(content), dtype=str)
            
            target_ids = [norm_id(eid) for eid in arguments.get("employee_ids", [])]
            
            # Find ID column (inlined normalization — no external dependency)
            id_col = next((c for c in df.columns if any(k in str(c).strip().lower() for k in ["employee id", "employee code", "associate id", "ee code", "employee_code"])), df.columns[0])
            
            # Filter
            filtered_df = df[df[id_col].apply(norm_id).isin(target_ids)]
            
            results = filtered_df.to_dict(orient="records")
            summary = save_results_to_excel(results, "Selective_Extraction")
            return [types.TextContent(type="text", text=json.dumps(summary, indent=2, default=_json_default))]

        elif name == "read_audit_report":
            path = arguments.get("file_path", "").strip().strip('"')
            sheet = arguments.get("sheet_name")
            import pandas as pd
            
            # Load all sheets if no specific sheet requested
            if not sheet:
                xls = pd.ExcelFile(path)
                result = {}
                for s in xls.sheet_names:
                    df = pd.read_excel(xls, sheet_name=s)
                    result[s] = df.to_dict(orient="records")
            else:
                df = pd.read_excel(path, sheet_name=sheet) if path.lower().endswith(('.xlsx', '.xls')) else pd.read_csv(path)
                result = df.to_dict(orient="records")
                
            return [types.TextContent(type="text", text=json.dumps(result, indent=2, default=_json_default))]

        elif name == "get_file_schema":
            path = arguments.get("file_path").strip().strip('"')
            sheet = arguments.get("sheet_name")
            df, error = _get_cached_df(path, sheet)
            if error: return [types.TextContent(type="text", text=error)]
            
            try:
                import re
                date_cols = []
                for col in df.columns:
                    if df[col].dtype == 'object':
                        # Check sample for date-like pattern (YYYY-MM-DD or MM/DD/YYYY)
                        sample = df[col].dropna().head(5).astype(str).tolist()
                        if any(re.search(r'\d{4}-\d{2}-\d{2}|\d{2}/\d{2}/\d{4}', s) for s in sample):
                            date_cols.append(col)

                schema = {
                    "filename": os.path.basename(path),
                    "sheet_name": sheet or "Auto-detected",
                    "total_rows": len(df),
                    "columns": [str(c) for c in df.columns],
                    "dtypes": {str(c): str(t) for c, t in df.dtypes.items()},
                    "date_hint_columns": date_cols,
                    "sql_note": f"If querying columns {date_cols}, use CAST(col AS DATE) for proper comparison." if date_cols else None,
                    "sample_data": df.head(5).to_dict(orient="records")
                }
                return [types.TextContent(type="text", text=json.dumps(schema, indent=2, default=_json_default))]
            except Exception as e:
                return [types.TextContent(type="text", text=f"Error analyzing schema: {str(e)}")]

        elif name == "query_data_sql":
            path = arguments.get("file_path").strip().strip('"')
            sheet = arguments.get("sheet_name")
            sql = arguments.get("sql_query")
            
            df, error = _get_cached_df(path, sheet)
            if error: return [types.TextContent(type="text", text=error)]
            
            try:
                con = duckdb.connect(database=':memory:')
                con.register('data', df)
                
                res_df = con.execute(sql).df()
                
                result = {
                    "query": sql,
                    "rows_found": len(res_df),
                    "data": res_df.to_dict(orient="records") if len(res_df) < 500 else res_df.head(100).to_dict(orient="records"),
                    "note": "Full data returned." if len(res_df) < 500 else "Data truncated (first 100 rows). Use more specific SQL filters."
                }
                return [types.TextContent(type="text", text=json.dumps(result, indent=2, default=_json_default))]
            except Exception as e:
                return [types.TextContent(type="text", text=f"SQL Error: {str(e)}")]

        raise ValueError(f"Unknown tool: {name}")

    except Exception as e:
        import traceback
        return [types.TextContent(type="text", text=f"Error in '{name}': {e}\n\n{traceback.format_exc()}")]


# ── SSE transport (for Vercel / HTTP) ────────────────────────────────────────
sse = SseServerTransport("/messages")

async def handle_sse(request):
    async with sse.connect_sse(request.scope, request.receive, request._send) as streams:
        await server.run(
            streams[0], streams[1],
            InitializationOptions(
                server_name="audit-tool-server",
                server_version="0.1.0",
                capabilities=server.get_capabilities(
                    notification_options=NotificationOptions(),
                    experimental_capabilities={},
                ),
            ),
        )

mcp_app = Starlette(routes=[
    Route("/sse", endpoint=handle_sse),
    Mount("/messages", app=sse.handle_post_message),
])

# ── Stdio transport (for local Claude Desktop) ────────────────────────────────
import asyncio
from mcp.server.stdio import stdio_server

async def run_stdio():
    async with stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream, write_stream,
            InitializationOptions(
                server_name="audit-tool-server",
                server_version="0.1.0",
                capabilities=server.get_capabilities(
                    notification_options=NotificationOptions(),
                    experimental_capabilities={},
                ),
            ),
        )

if __name__ == "__main__":
    asyncio.run(run_stdio())
