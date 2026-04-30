import pandas as pd
import io
import re
import numpy as np
from datetime import datetime, date

# --- Constants from the original utils ---
UZIO_RAW_MAPPING = {
    'Employee ID*': 'Employee ID',
    'Employee First Name*': 'First Name',
    'Employee Last Name*': 'Last Name',
    'Employee Middle Initial': 'Middle Initial',
    'Employee Suffix': 'Suffix',
    'Employment Status*': 'Employment Status',
    'Date of Hire*': 'Hire Date',
    'Original DOH': 'Original Hire Date',
    'Termination Date': 'Termination Date',
    'Termination Reason': 'Termination Reason',
    'Employment Type*': 'Employment Type',
    'Pay Type*': 'Pay Type',
    'Annual Salary(Digits)**': 'Annual Salary',
    'Hourly Pay Rate**': 'Hourly Pay Rate',
    'Working Hours per Week(Digits)**': 'Working Hours',
    'Job Title': 'Job Title',
    'Department': 'Department',
    'Official Email*': 'Work Email',
    'Personal Email': 'Personal Email',
    'Phone Number(Digits)': 'Phone Number',
    'Employee SSN': 'SSN',
    'Employee Date of Birth*': 'DOB',
    'Employee Gender*': 'Gender',
    'Employee Tobacco usage in last 12 months': 'Tobacco User',
    'FLSA Classification': 'FLSA Classification',
    'Employee Address Line 1': 'Address Line 1',
    'Employee Address Line 2': 'Address Line 2',
    'City*': 'City',
    'Zipcode*': 'Zip',
    'State(Abbreviation)*': 'State',
    'Mailing Address Line 1': 'Mailing Address Line 1',
    'Mailing Address Line 2': 'Mailing Address Line 2',
    'Mailing City': 'Mailing City',
    'Mailing Zipcode': 'Mailing Zip',
    'Mailing State(Abbreviation)': 'Mailing State',
    'Reporting Manager ID': 'Reports To ID',
    'Work Location': 'Work Location',
    'License Number*': 'License Number',
    'License Expiration Date': 'License Expiration Date'
}

HOURLY_ONLY_JOB_TITLES = {
    "driver", "lead driver", "walker", "helper", "driver-lite", "driver-step van",
    "driver-unscheduled", "ddu dedicated", "ddu shared", "delivery associate",
    "delivery associates", "driver -major appliance"
}

# --- Core Utility Functions ---

def clean_money_val(x):
    if pd.isna(x) or x == "": return 0.0
    s = str(x).strip().replace("$", "").replace("%", "").replace(",", "")
    s = s.replace("(", "-").replace(")", "")
    try: return float(s)
    except: return 0.0

def smart_read_df(content, filename="", sheet_name=None, header='infer', required_columns=None, fallback_columns=None, **kwargs):
    """
    Robustly reads Excel or CSV from bytes. 
    If sheet_name is provided and it's an Excel file, it reads that sheet.
    If required_columns is provided, it scans for the header row.
    Falls back to CSV if Excel reading fails or if the extension is .csv.
    """
    import io
    file_io = io.BytesIO(content)
    is_csv = str(filename).lower().endswith('.csv')
    
    if not is_csv:
        # Try Excel
        try:
            xls = pd.ExcelFile(file_io)
            # Use provided sheet or search all sheets
            sheets = [sheet_name] if sheet_name and sheet_name in xls.sheet_names else xls.sheet_names
            
            for sheet in sheets:
                if required_columns:
                    df_peek = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=50)
                    header_row_idx = None
                    for idx, row in df_peek.iterrows():
                        row_vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
                        if all(any(col.lower() in v for v in row_vals) for col in required_columns):
                            header_row_idx = idx
                            break
                    
                    if header_row_idx is not None:
                        return pd.read_excel(xls, sheet_name=sheet, header=header_row_idx, **kwargs)
                    
                    # Try fallback columns
                    if fallback_columns:
                        for idx, row in df_peek.iterrows():
                            row_vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
                            if all(any(col.lower() in v for v in row_vals) for col in fallback_columns):
                                return pd.read_excel(xls, sheet_name=sheet, header=idx, **kwargs)
                else:
                    return pd.read_excel(xls, sheet_name=sheet, header=header, **kwargs)
        except Exception:
            pass

    # Try CSV
    try:
        file_io.seek(0)
        if required_columns:
            # We need to peek to find the header
            wrapper = io.TextIOWrapper(file_io, encoding='utf-8', errors='replace')
            header_row_idx = None
            for i, line in enumerate(wrapper):
                line_lower = line.lower()
                if all(col.lower() in line_lower for col in required_columns):
                    header_row_idx = i
                    break
                if i > 100: break
            
            if header_row_idx is not None:
                file_io.seek(0)
                return pd.read_csv(file_io, header=header_row_idx, **kwargs)
            
            if fallback_columns:
                file_io.seek(0)
                wrapper = io.TextIOWrapper(file_io, encoding='utf-8', errors='replace')
                for i, line in enumerate(wrapper):
                    line_lower = line.lower()
                    if all(col.lower() in line_lower for col in fallback_columns):
                        header_row_idx = i
                        break
                    if i > 100: break
                if header_row_idx is not None:
                    file_io.seek(0)
                    return pd.read_csv(file_io, header=header_row_idx, **kwargs)

        file_io.seek(0)
        return pd.read_csv(file_io, header=header, **kwargs)
    except Exception:
        return pd.DataFrame()


def norm_colname(c: str) -> str:
    if c is None: return ""
    c = str(c).replace("\n", " ").replace("\r", " ").replace("\u00A0", " ")
    c = c.replace("’", "'").replace("“", '"').replace("”", '"')
    c = re.sub(r'\(.*?\)', '', c)
    c = re.sub(r"\s+", " ", c).strip().replace("*", "")
    return c.strip('"').strip("'")

def norm_blank(x):
    if x is None or (isinstance(x, float) and np.isnan(x)): return ""
    if isinstance(x, str) and x.strip().lower() in {"", "nan", "none", "null"}: return ""
    return x

def normalize_id(x):
    if pd.isna(x) or x is None: return ""
    s = str(x).strip()
    if s.endswith(".0"): s = s[:-2]
    return s.lstrip("0") if s != "0" else "0"

def try_parse_date(x):
    x = norm_blank(x)
    if x == "": return ""
    try:
        ts = pd.to_datetime(x, errors='coerce')
        return ts.strftime("%m/%d/%Y") if pd.notna(ts) else str(x)
    except: return str(x)

def format_pay_date(date_val):
    if pd.isna(date_val) or str(date_val).strip() in ["", "nan", "NaT"]: return "Unknown"
    try:
        dt = pd.to_datetime(date_val)
        return dt.strftime('%Y-%m-%d')
    except: return str(date_val).strip()

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    norm_cols = [norm_colname(c).casefold() for c in df.columns]
    seen, to_keep = set(), []
    for i, nc in enumerate(norm_cols):
        if nc not in seen:
            seen.add(nc)
            to_keep.append(i)
    return df.iloc[:, to_keep]

def safe_val(df, idx, col):
    if idx is None or col not in df.columns: return ""
    val = df.loc[idx, col]
    return val.iloc[0] if isinstance(val, pd.Series) else val

def norm_ssn_canonical(x):
    x = norm_blank(x)
    if x == "": return ""
    s = re.sub(r"\D", "", str(x).strip().replace("-", "").replace(" ", "").replace(".0", ""))
    return s.zfill(9)[-9:] if s else ""

def find_header_and_data(file_content, filename):
    file_io = io.BytesIO(file_content)
    if filename.lower().endswith('.csv'):
        df_peek = pd.read_csv(file_io, header=None, nrows=50)
        header_idx = 0
        for i, row in df_peek.iterrows():
            row_str = " ".join([str(x).lower() for x in row if pd.notna(x)])
            if any(k in row_str for k in ["employee id", "employee name", "associate id"]):
                header_idx = i
                break
        file_io.seek(0)
        df = pd.read_csv(file_io, header=header_idx)
        header_top = df_peek.iloc[header_idx - 1].tolist() if header_idx > 0 else None
    else:
        xls = pd.ExcelFile(file_io)
        sheet = xls.sheet_names[1] if len(xls.sheet_names) > 1 and "criteria" in xls.sheet_names[0].lower() else xls.sheet_names[0]
        df_peek = pd.read_excel(xls, sheet_name=sheet, header=None, nrows=50)
        header_idx = 0
        for i, row in df_peek.iterrows():
            row_str = " ".join([str(x).lower() for x in row if pd.notna(x)])
            if any(k in row_str for k in ["employee id", "employee name", "associate id"]):
                header_idx = i
                break
        df = pd.read_excel(xls, sheet_name=sheet, header=header_idx)
        header_top = df_peek.iloc[header_idx - 1].tolist() if header_idx > 0 else None
    return df, header_top, filename

def read_uzio_raw_file(content):
    """Reads Uzio raw census file, supporting both Excel (with sheet logic) and CSV."""
    try:
        # Try Excel with specific sheet and header row
        df = smart_read_df(content, sheet_name='Employee Details', header=3)
        if df.empty:
            # Try CSV fallback (no sheet name)
            df = smart_read_df(content, header=3)
            
        if df.empty:
            # Try without header skip if still empty
            df = smart_read_df(content)

        df.columns = [str(c).strip() for c in df.columns]
        norm_mapping = {norm_colname(k).casefold(): v for k, v in UZIO_RAW_MAPPING.items()}
        df.columns = [norm_mapping.get(norm_colname(c).casefold(), c) for c in df.columns]
        
        if 'Employee ID' in df.columns:
            df['Employee ID'] = df['Employee ID'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
            
        return df
    except Exception as e:
        import sys
        sys.stderr.write(f"Error reading Uzio Raw File: {e}\n")
        return pd.DataFrame()

def is_hourly_only_job_title(jt_val: str) -> bool:
    jt = jt_val.strip().lower()
    return jt in HOURLY_ONLY_JOB_TITLES or jt.endswith("driver")

def get_identity_match_map(df_uzio, df_vendor, uzio_id_col, vendor_id_col, uzio_ssn_col, vendor_ssn_col):
    """Maps Uzio employee IDs to vendor IDs via SSN matching."""
    uz_ssn_map = {norm_ssn_canonical(row[uzio_ssn_col]): str(row[uzio_id_col]).strip()
                  for _, row in df_uzio.iterrows() if norm_ssn_canonical(row.get(uzio_ssn_col, ""))}
    result = {}
    for _, row in df_vendor.iterrows():
        ssn = norm_ssn_canonical(row.get(vendor_ssn_col, ""))
        vid = str(row[vendor_id_col]).strip()
        if ssn and ssn in uz_ssn_map:
            result[uz_ssn_map[ssn]] = vid
    return result

def norm_id(x):
    if pd.isna(x): return ""
    s = str(x).strip()
    if s.endswith(".0"): s = s[:-2]
    return s.lstrip("0")

def normalize_space_and_case(x):
    if not x: return ""
    return re.sub(r"\s+", " ", str(x).strip()).lower()

def as_float_or_none(x):
    try: return float(str(x).replace(",", "").replace("$", "").strip())
    except: return None

def detect_duplicate_ssns(df, ssn_col):
    """Identifies duplicate SSNs in a dataframe."""
    df = df.copy()
    df['__norm_ssn'] = df[ssn_col].apply(norm_ssn_canonical)
    dupes = df[df['__norm_ssn'] != ""][df.duplicated(subset=['__norm_ssn'], keep=False)]
    return dupes

def normalize_paytype_text(x):
    return str(norm_blank(x)).strip().lower()

def paytype_bucket(pt_norm: str) -> str:
    if "salary" in pt_norm or "salaried" in pt_norm: return "salaried"
    if "hourly" in pt_norm or "hour" in pt_norm: return "hourly"
    return "other"

def normalize_reason_text(x):
    s = str(norm_blank(x)).strip().replace("\u00A0", " ")
    s = s.replace("’", "'").replace("“", '"').replace("”", '"')
    s = re.sub(r"\s+", " ", s).strip()
    return s.strip('"').strip("'").lower()

def is_termination_reason_field(f: str) -> bool:
    return "termination reason" in norm_colname(f).lower()

def is_employment_status_field(f: str) -> bool:
    return any(k in norm_colname(f).lower() for k in ["employment status", "position status", "worker status"])

def status_contains_any(s: str, needles: list) -> bool:
    s = str(s).lower()
    return any(n in s for n in needles)

def uzio_is_active(s: str) -> bool:
    return "active" in str(s).lower()

def uzio_is_terminated(s: str) -> bool:
    return any(x in str(s).lower() for x in ["terminated", "inactive", "quit", "resign"])

ALLOWED_TERM_REASONS = {
    "quit without notice", "no reason given", "misconduct", "abandoned job",
    "advancement (better job with higher pay)", "no-show (never started employment)",
    "performance", "personal", "scheduling conflicts (schedules don't work)",
    "attendance", "reduction in force", "reorganization", "mutual agreement",
    "import created action", "advancement", "no-show", "management", "layoff"
}


# ─────────────────────────────────────────────────────────────────────────────
# Uzio Census Template Generator (ported from utils/audit_utils.py)
# ─────────────────────────────────────────────────────────────────────────────

def generate_uzio_template(df_source, vendor_field_map, fix_options=None):
    """
    Generate an Uzio Census Template DataFrame from a source DataFrame.

    df_source              : pandas DataFrame with vendor column names (already normalized
                             by the caller's preprocess step).
    vendor_field_map       : dict mapping STANDARD field name (e.g. 'Employee ID') to the
                             actual column name in df_source.
    fix_options            : optional dict of opt-in auto-fix toggles. Recognised keys:
                             fix_status, fix_inactive, fix_zip, fix_type, fix_emails,
                             fix_position, fix_dol_status, fix_license, fix_flsa.
    """
    uzio_headers = list(UZIO_RAW_MAPPING.keys())
    df_uzio = pd.DataFrame(columns=uzio_headers)

    for uzio_header, std_name in UZIO_RAW_MAPPING.items():
        # These three are populated by the caller via the Job/Location mapping UI;
        # leave blank here so the caller can write the user-edited values.
        if std_name in ['Job Title', 'Department', 'Work Location']:
            df_uzio[uzio_header] = ""
            continue

        vendor_col = vendor_field_map.get(std_name)
        if vendor_col and vendor_col in df_source.columns:
            series = df_source[vendor_col].copy()

            if std_name == 'Middle Initial':
                series = series.apply(lambda x: str(x).strip()[0] if pd.notna(x) and str(x).strip() else "")
            elif std_name in ['Hire Date', 'Original Hire Date', 'Termination Date', 'DOB']:
                def format_date(d):
                    if pd.isna(d) or str(d).strip() == "": return ""
                    try:
                        dt = pd.to_datetime(str(d).strip(), errors='coerce')
                        if pd.isna(dt): return str(d).strip()
                        return dt.strftime('%d/%m/%Y')
                    except Exception:
                        return str(d).strip()
                series = series.apply(format_date)
            elif std_name == 'License Expiration Date':
                def format_license_exp_date(d):
                    if pd.isna(d) or str(d).strip() == "": return ""
                    d_str = str(d).strip()
                    if '00/00/0000' in d_str or d_str in ('0', '00', '0000'): return ""
                    try:
                        dt = pd.to_datetime(d_str, errors='coerce')
                        if pd.isna(dt): return ""
                        return dt.strftime('%m/%d/%Y')
                    except Exception:
                        return ""
                series = series.apply(format_license_exp_date)
            elif std_name == 'SSN':
                series = series.apply(lambda x: str(x).replace("-", "").strip() if pd.notna(x) else "")
            elif std_name == 'Gender':
                def format_gender(g):
                    if pd.isna(g) or str(g).strip() == "": return ""
                    g_str = str(g).strip().lower()
                    if g_str.startswith('m'): return "Male"
                    if g_str.startswith('f'): return "Female"
                    return ""
                series = series.apply(format_gender)
            elif std_name == 'Employment Status':
                def format_status(row):
                    x = row[vendor_col]
                    if pd.isna(x): return ""
                    s = str(x).strip().lower()
                    if not s: return ""

                    if fix_options and fix_options.get('fix_status', False):
                        if 'not hired' in s: return 'EXCLUDE'
                        if 'leave' in s: return 'ACTIVE'
                        if 'term' in s: return 'TERMINATED'
                        if 'active' in s: return 'ACTIVE'

                    if fix_options and fix_options.get('fix_inactive', False):
                        if 'inactive' in s:
                            term_col = vendor_field_map.get('Termination Date')
                            if term_col and pd.notna(row.get(term_col)) and str(row.get(term_col)).strip() != "":
                                return 'TERMINATED'
                            return 'ACTIVE'
                    elif 'inactive' in s:
                        return 'INACTIVE'

                    return str(x).strip().upper()
                series = df_source.apply(format_status, axis=1)
            elif std_name in ['Zip', 'Mailing Zip']:
                def format_zip(z):
                    if pd.isna(z) or str(z).strip() == "": return ""
                    if fix_options and fix_options.get('fix_zip', False):
                        s = str(z).split('.')[0].split('-')[0]
                        z_clean = re.sub(r'\D', '', s.strip())
                        if not z_clean: return ""
                        if len(z_clean) == 4:
                            return '0' + z_clean
                        return z_clean[:5]
                    return str(z).strip()
                series = series.apply(format_zip)
            elif std_name == 'Employment Type':
                def format_emp_type(et):
                    if pd.isna(et) or str(et).strip() == "": return ""
                    et_str = str(et).strip().lower()
                    if fix_options and fix_options.get('fix_type', False):
                        if 'full' in et_str: return 'Full Time'
                        if 'part' in et_str: return 'Part Time'
                        if 'season' in et_str: return 'Seasonal'
                        if 'other' in et_str: return 'Other'
                        if 'intern' in et_str: return 'Part Time'
                    return str(et).strip()
                series = series.apply(format_emp_type)
            elif std_name == 'Termination Reason':
                def format_term_reason(tr):
                    if pd.isna(tr) or str(tr).strip() == "": return ""
                    tr_str = str(tr).strip().lower()
                    if "involuntary" in tr_str or "invluntary" in tr_str:
                        return "Involuntary Termination of Employment"
                    if "voluntary" in tr_str or "quit" in tr_str:
                        return "Voluntary Termination of Employment"
                    if "death" in tr_str: return "Death"
                    if "retire" in tr_str: return "Retirement"
                    if "disability" in tr_str: return "Permanent Disability"
                    if "transfer" in tr_str: return "Transfer"
                    return "Other"
                series = series.apply(format_term_reason)

            df_uzio[uzio_header] = series
        else:
            df_uzio[uzio_header] = ""

    fix_logs = []

    if 'Employment Status*' in df_uzio.columns:
        df_uzio = df_uzio[df_uzio['Employment Status*'] != 'EXCLUDE'].copy()

    emp_ids = df_uzio['Employee ID*'] if 'Employee ID*' in df_uzio.columns else df_uzio.index

    # Work-email fallback to personal email
    if fix_options and fix_options.get('fix_emails', False):
        if 'Official Email*' in df_uzio.columns and 'Personal Email' in df_uzio.columns:
            missing_work_mask = df_uzio['Official Email*'].isna() | (df_uzio['Official Email*'].astype(str).str.strip() == "")
            has_personal_mask = df_uzio['Personal Email'].notna() & (df_uzio['Personal Email'].astype(str).str.strip() != "")
            combined_mask = missing_work_mask & has_personal_mask
            for idx in df_uzio[combined_mask].index:
                fix_logs.append({
                    "Employee": emp_ids[idx], "Field Fixed": "Official Email*",
                    "Original Value": "(Blank)", "New Value": df_uzio.loc[idx, 'Personal Email'],
                    "Fix Applied": "Fallback to Personal Email"
                })
            df_uzio.loc[combined_mask, 'Official Email*'] = df_uzio.loc[combined_mask, 'Personal Email']

    # Position auto-fill from Department description (Paycom)
    if fix_options and fix_options.get('fix_position', False):
        if 'Job Title' in df_uzio.columns:
            dept_desc_col = next(
                (c for c in df_source.columns
                 if str(c).lower().strip().replace(' ', '_') == 'department_desc'
                 or str(c).lower().strip() == 'department_description'),
                None
            )
            if dept_desc_col:
                missing_job_mask = df_uzio['Job Title'].isna() | (df_uzio['Job Title'].astype(str).str.strip() == "")
                has_dept_mask = df_source[dept_desc_col].notna() & (df_source[dept_desc_col].astype(str).str.strip() != "")
                combined_mask = missing_job_mask & has_dept_mask
                for idx in df_uzio[combined_mask].index:
                    fix_logs.append({
                        "Employee": emp_ids[idx], "Field Fixed": "Job Title",
                        "Original Value": "(Blank)", "New Value": df_source.loc[idx, dept_desc_col],
                        "Fix Applied": "Fallback to Department"
                    })
                df_uzio.loc[combined_mask, 'Job Title'] = df_source.loc[combined_mask, dept_desc_col]

    # DOL_Status auto-fill (Paycom)
    if fix_options and fix_options.get('fix_dol_status', False):
        dol_col = None
        for cand in ['dol_status', 'dol status', 'worker category description']:
            cand_col = next((c for c in df_source.columns if str(c).lower().strip().replace('_', ' ') == cand), None)
            if cand_col:
                dol_col = cand_col
                break
        if dol_col and 'Employment Type*' in df_uzio.columns:
            blank_dol_mask = df_source[dol_col].isna() | (df_source[dol_col].astype(str).str.strip() == "")
            for idx in df_uzio[blank_dol_mask].index:
                fix_logs.append({
                    "Employee": emp_ids[idx], "Field Fixed": "Employment Type*",
                    "Original Value": "(Blank)", "New Value": "Full Time",
                    "Fix Applied": "Default blank to Full Time"
                })
            df_uzio.loc[blank_dol_mask, 'Employment Type*'] = "Full Time"

    # License rules
    if fix_options and fix_options.get('fix_license', False):
        lic_num_col = 'License Number*'
        lic_exp_col = 'License Expiration Date'
        if lic_exp_col in df_uzio.columns:
            bad_exp_mask = df_uzio[lic_exp_col].astype(str).str.strip().isin(
                ['00/00/0000', '0', '00', '0000', 'nan', 'NaT', '']
            )
            for idx in df_uzio[bad_exp_mask].index:
                fix_logs.append({
                    "Employee": emp_ids[idx], "Field Fixed": "License Expiration Date",
                    "Original Value": df_uzio.loc[idx, lic_exp_col], "New Value": "(Blank)",
                    "Fix Applied": "Cleared Invalid Date Placeholder"
                })
            df_uzio.loc[bad_exp_mask, lic_exp_col] = ""
            if lic_num_col in df_uzio.columns:
                no_license_mask = df_uzio[lic_num_col].isna() | (df_uzio[lic_num_col].astype(str).str.strip() == "") | (df_uzio[lic_num_col].astype(str).str.strip() == 'nan')
                exp_not_blank = df_uzio[lic_exp_col].astype(str).str.strip() != ""
                combined_mask = no_license_mask & exp_not_blank
                for idx in df_uzio[combined_mask].index:
                    fix_logs.append({
                        "Employee": emp_ids[idx], "Field Fixed": "License Expiration Date",
                        "Original Value": df_uzio.loc[idx, lic_exp_col], "New Value": "(Blank)",
                        "Fix Applied": "Cleared Date due to missing License Number"
                    })
                df_uzio.loc[no_license_mask, lic_exp_col] = ""

    # Pay-Type / FLSA rules
    if 'Pay Type*' in df_uzio.columns:
        # Driver special case: force Hourly + Non-Exempt
        if 'Job Title' in df_uzio.columns:
            driver_mask = df_uzio['Job Title'].astype(str).str.lower().str.contains('driver', na=False)
            pt_to_fix = driver_mask & ((df_uzio['Pay Type*'].astype(str).str.lower().str.strip() != 'hourly') | df_uzio['Pay Type*'].isna() | (df_uzio['Pay Type*'] == ""))
            for idx in df_uzio[pt_to_fix].index:
                cur = df_uzio.loc[idx, 'Pay Type*']
                fix_logs.append({
                    "Employee": emp_ids[idx], "Field Fixed": "Pay Type*",
                    "Original Value": cur if pd.notna(cur) and str(cur).strip() else "(Blank)",
                    "New Value": "Hourly", "Fix Applied": "Forced Hourly for Driver Position"
                })
            df_uzio.loc[driver_mask, 'Pay Type*'] = "Hourly"

            if 'FLSA Classification' in df_uzio.columns:
                flsa_to_fix = driver_mask & ((df_uzio['FLSA Classification'].astype(str).str.lower().str.strip() != 'non-exempt') | df_uzio['FLSA Classification'].isna() | (df_uzio['FLSA Classification'] == ""))
                for idx in df_uzio[flsa_to_fix].index:
                    cur = df_uzio.loc[idx, 'FLSA Classification']
                    fix_logs.append({
                        "Employee": emp_ids[idx], "Field Fixed": "FLSA Classification",
                        "Original Value": cur if pd.notna(cur) and str(cur).strip() else "(Blank)",
                        "New Value": "Non-Exempt", "Fix Applied": "Forced Non-Exempt for Driver Position"
                    })
                df_uzio.loc[driver_mask, 'FLSA Classification'] = "Non-Exempt"

        pay_type_series = df_uzio['Pay Type*'].astype(str).str.lower().str.strip()

        hourly_mask = pay_type_series.str.contains('hour', na=False)
        df_uzio.loc[hourly_mask, 'Pay Type*'] = "Hourly"
        if 'Annual Salary(Digits)**' in df_uzio.columns:
            df_uzio.loc[hourly_mask, 'Annual Salary(Digits)**'] = ""
            if fix_options and fix_options.get('fix_flsa', False):
                if 'FLSA Classification' in df_uzio.columns:
                    df_uzio.loc[hourly_mask, 'FLSA Classification'] = "Non-Exempt"

        salary_mask = pay_type_series.str.contains('salar', na=False)
        df_uzio.loc[salary_mask, 'Pay Type*'] = "Salaried"
        if 'Hourly Pay Rate**' in df_uzio.columns:
            # Use string "0" — modern pandas refuses int writes into string-dtype columns.
            df_uzio.loc[salary_mask, 'Hourly Pay Rate**'] = "0"
        if 'Working Hours per Week(Digits)**' in df_uzio.columns:
            df_uzio.loc[salary_mask, 'Working Hours per Week(Digits)**'] = ""

        if fix_options and fix_options.get('fix_flsa', False):
            if 'FLSA Classification' in df_uzio.columns:
                df_uzio.loc[salary_mask, 'FLSA Classification'] = "Exempt"
                blank_flsa_mask = df_uzio['FLSA Classification'].isna() | (df_uzio['FLSA Classification'].astype(str).str.strip() == "")
                df_uzio.loc[blank_flsa_mask, 'FLSA Classification'] = "Non-Exempt"

    df_uzio.attrs['fix_logs'] = pd.DataFrame(fix_logs) if fix_logs else pd.DataFrame(
        columns=["Employee", "Field Fixed", "Original Value", "New Value", "Fix Applied"]
    )
    return df_uzio


def inject_into_uzio_template(df_uzio, template_path):
    """
    Inject a generated Uzio DataFrame into the standard Uzio .xlsm template,
    preserving every other sheet, instructions, and the VBA project.

    template_path may be a path string or a file-like object.
    """
    import openpyxl
    import os

    if isinstance(template_path, str):
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found at {template_path}")

    wb = openpyxl.load_workbook(template_path, keep_vba=True)
    ws = wb['Employee Details']

    header_row = 4  # fallback
    for r in range(1, 10):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and re.sub(r'\s+', ' ', str(val)).strip() == 'Employee First Name*':
                header_row = r
                break
        if header_row == r:
            break

    headers_in_template = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val:
            norm_val = re.sub(r'\s+', ' ', str(val)).strip()
            headers_in_template[norm_val] = col_idx

    start_row = header_row + 1
    for row_idx, row_data in df_uzio.iterrows():
        excel_row = start_row + row_idx
        for col_name in df_uzio.columns:
            c_name_strip = re.sub(r'\s+', ' ', str(col_name)).strip()
            if c_name_strip in headers_in_template:
                col_idx = headers_in_template[c_name_strip]
                val = row_data[col_name]
                if pd.notna(val) and val != "":
                    ws.cell(row=excel_row, column=col_idx, value=val)

    return wb


def resolve_uzio_template_path():
    """
    Locate the Uzio_Census_Template.xlsm. Looks in:
      1. <audit_fast_api>/templates/
      2. <audit_fast_api>/../templates/   (parent Streamlit project)
      3. CWD/templates/
    Returns the first existing path or None.
    """
    import os
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # audit_fast_api/
    candidates = [
        os.path.join(here, 'templates', 'Uzio_Census_Template.xlsm'),
        os.path.join(os.path.dirname(here), 'templates', 'Uzio_Census_Template.xlsm'),
        os.path.join(os.getcwd(), 'templates', 'Uzio_Census_Template.xlsm'),
    ]
    for p in candidates:
        if os.path.isfile(p):
            return p
    return None
