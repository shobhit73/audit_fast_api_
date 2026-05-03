"""File-shape guards for vendor-specific tools.

The MCP server exposes ~14 ADP-only tools, ~14 Paycom-only tools, and a few
that take BOTH a UZIO file and a vendor file. When a tool takes a SINGLE file
(e.g. adp_prior_payroll_sanity, paycom_census_sanity), there's nothing in the
parameter shape that prevents an agent from passing the wrong-vendor file.

This module sniffs a file's columns / sheet names to detect whether it's an
ADP, Paycom, or UZIO export, and raises a clear ValueError when the wrong
vendor's file is handed to a vendor-specific tool. Pair every single-file
vendor tool with a `require_vendor(content, filename, "adp", "<tool_name>")`
call at the top of its handler.
"""

from __future__ import annotations
import io


# Lowercase substrings that, if present in column names or sheet contents,
# strongly suggest the file is from this vendor. Keep these lists conservative
# (high precision, low recall is fine here -- two matches outweigh false alarms).
ADP_COLUMN_MARKERS = {
    "associate id", "associate_id",
    "position id", "position_id",
    "worked in state",
    "file number", "file #",
    "period beginning date", "period ending date",
    "additional earnings  :",  # ADP's distinctive double-space-colon pattern
    "voluntary deduction :",
}

PAYCOM_COLUMN_MARKERS = {
    "employee_code",
    "ss_number",
    "dol_status",
    "department_desc",
    "department_code",
    "exempt_status",
    "type_code",
    "type_description",
    "code_description",
    "legal_firstname",
    "legal_lastname",
}

# UZIO uses pipe-delimited section|field column names (e.g. "Personal|SSN")
# and a distinctive 3-row preamble in raw exports.
UZIO_COLUMN_MARKERS = {
    "personal|ssn",
    "job|employee id",
    "job|department",
    "job|location",
    "job|job title",
    "job|hire date",
    "personal|first name",
    "personal|last name",
    "compensation|annual salary",
}

# ADP exports often arrive as xlsx workbooks with a 'Report Criteria' preamble
# sheet plus the actual data sheet (named after the report itself). These
# sheet names are highly distinctive of ADP's reporting platform.
ADP_SHEET_MARKERS = {
    "report criteria",
    "prior payroll register",
    "payroll register",
    "earnings register",
    "deduction register",
    "tax register",
}

# Note: 'employee details' is intentionally NOT a UZIO marker -- ADP's RAW
# xlsx exports use 'EMPLOYEE DETAILS' as a row-1 section header inside their
# data sheet. UZIO is detected by its pipe-delimited column convention
# ('Personal|SSN', 'Job|Employee ID', ...) which is unique to UZIO.
UZIO_SHEET_MARKERS: set[str] = set()


def _sniff(content: bytes, filename: str):
    """Return (sheet_names_lower, column_text_pieces_lower). Cheap: reads only
    the first sheet's first 10 rows. Falls back to empty lists on parse error.
    """
    name = (filename or "").lower()
    sheets: list[str] = []
    cell_pieces: list[str] = []

    try:
        if name.endswith(".csv"):
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content), nrows=10, header=None, dtype=str)
            for _, row in df.iterrows():
                for v in row.dropna().tolist():
                    s = str(v).strip().lower()
                    if s:
                        cell_pieces.append(s)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheets = [s.lower() for s in wb.sheetnames]
            # Scan every sheet's first 10 rows. ADP exports often have a small
            # 'Report Criteria' preamble sheet first and the actual data on a
            # later sheet ('Prior Payroll Register', 'Employee Details', etc.),
            # so we must look beyond just sheet 0.
            for sheet_name in wb.sheetnames[:6]:  # cap at 6 sheets for safety
                ws = wb[sheet_name]
                for r in range(1, min(ws.max_row or 0, 10) + 1):
                    for c in range(1, min(ws.max_column or 0, 250) + 1):
                        v = ws.cell(r, c).value
                        if v is not None:
                            s = str(v).strip().lower()
                            if s:
                                cell_pieces.append(s)
    except Exception:
        return [], []
    return sheets, cell_pieces


def detect_vendor(content: bytes, filename: str) -> dict:
    """Detect a file's vendor. Returns {vendor: 'adp'|'paycom'|'uzio'|'unknown',
    evidence: [list of markers found]}.
    """
    sheets, pieces = _sniff(content, filename)
    if not pieces and not sheets:
        return {"vendor": "unknown", "evidence": ["could not parse file"]}

    sheet_set = set(sheets)

    # ADP sheet-name check first - 'Report Criteria', 'Prior Payroll Register'
    # etc. are highly distinctive of ADP exports.
    adp_sheet_hits = sorted(ADP_SHEET_MARKERS & sheet_set)
    if adp_sheet_hits:
        return {"vendor": "adp", "evidence": adp_sheet_hits}

    # UZIO check - pipe-delimited column markers (unique to UZIO).
    uzio_evidence = sorted(
        UZIO_SHEET_MARKERS.intersection(sheet_set)
        | {m for m in UZIO_COLUMN_MARKERS if any(m in p for p in pieces)}
    )
    if uzio_evidence:
        return {"vendor": "uzio", "evidence": uzio_evidence}

    piece_set = set(pieces)
    adp_hits = sorted({m for m in ADP_COLUMN_MARKERS if any(m in p for p in pieces)})
    paycom_hits = sorted({m for m in PAYCOM_COLUMN_MARKERS if m in piece_set})

    # Decisive single markers
    if "associate id" in piece_set or "associate_id" in piece_set:
        return {"vendor": "adp", "evidence": adp_hits or ["associate id"]}
    if "employee_code" in piece_set:
        return {"vendor": "paycom", "evidence": paycom_hits or ["employee_code"]}

    # Fallback: 2+ markers wins
    if len(adp_hits) >= 2 and len(adp_hits) > len(paycom_hits):
        return {"vendor": "adp", "evidence": adp_hits}
    if len(paycom_hits) >= 2 and len(paycom_hits) > len(adp_hits):
        return {"vendor": "paycom", "evidence": paycom_hits}
    if adp_hits and not paycom_hits:
        return {"vendor": "adp", "evidence": adp_hits}
    if paycom_hits and not adp_hits:
        return {"vendor": "paycom", "evidence": paycom_hits}
    return {"vendor": "unknown", "evidence": adp_hits + paycom_hits}


def require_vendor(content: bytes, filename: str, expected: str, tool_name: str):
    """Raise ValueError with a clear, actionable message if the file isn't
    from the expected vendor. 'unknown' is allowed through (the tool's own
    header detection will catch it later, and we don't want to block legitimate
    exotic exports).
    """
    info = detect_vendor(content, filename)
    if info["vendor"] in (expected, "unknown"):
        return info
    wrong = info["vendor"]
    suggestions = {
        "adp": {
            "uzio": "Use a Uzio-* tool, or feed this UZIO file to the right slot of an audit (e.g. uzio_file_path).",
            "paycom": "Use the paycom_* equivalent (e.g. paycom_prior_payroll_sanity does NOT exist; the Paycom equivalent for census cleanup is 'paycom_census_sanity').",
        },
        "paycom": {
            "uzio": "Use a Uzio-* tool, or feed this UZIO file to the right slot of an audit (e.g. uzio_file_path).",
            "adp": "Use the adp_* equivalent of this tool.",
        },
    }
    hint = suggestions.get(expected, {}).get(wrong, "Confirm with the user which file they meant to upload.")
    raise ValueError(
        f"{tool_name} expects an {expected.upper()} file, but the supplied "
        f"file '{filename}' looks like a {wrong.upper()} file "
        f"(detected markers: {info['evidence']}). {hint}"
    )
